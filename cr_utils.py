import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta
from io import BytesIO
import xlsxwriter

# ==============================================================================
# --- CONSTANTS & SHARED FUNCTIONS ---
# ==============================================================================

PASTEL_COLORS = {
    'red': '#ff6961',
    'orange': '#ffb347',
    'green': '#77dd77',
    'blue': '#3498DB',
    'grey': '#808080',
    'target_line': 'deepskyblue',
    'optimal_line': 'darkblue',
    'purple': '#8A2BE2'
}

def _get_stable_mode(series: pd.Series) -> float:
    """
    Computes the statistical mode of a cycle time series.
    Rounds to 2dp before mode selection to collapse floating-point noise
    (e.g. 97.0999 and 97.1001 are treated as identical → mode = 97.10).
    Ported from run_rate_utils for consistency.
    """
    if series.empty:
        return 0.0
    rounded = series.round(2)
    modes = rounded.mode()
    return float(modes.iloc[0]) if not modes.empty else float(series.mean())


def format_seconds_to_dhm(total_seconds):
    """Converts total seconds into a 'Xd Yh Zm' or 'Xs' string."""
    if pd.isna(total_seconds) or total_seconds < 0: return "N/A"
    
    if total_seconds < 60:
         return f"{total_seconds:.1f}s"

    total_minutes = int(total_seconds / 60)
    days = total_minutes // (60 * 24)
    remaining_minutes = total_minutes % (60 * 24)
    hours = remaining_minutes // 60
    minutes = remaining_minutes % 60
    parts = []
    if days > 0: parts.append(f"{days}d")
    if hours > 0: parts.append(f"{hours}h")
    if minutes > 0 or not parts: parts.append(f"{minutes}m")
    return " ".join(parts) if parts else "0m"

def load_logistics_plan(file):
    """Loads logistics plan CSV/Excel to extract PO information."""
    if not file: return pd.DataFrame()
    try:
        df = pd.read_csv(file) if file.name.endswith('.csv') else pd.read_excel(file)
        col_map = {col.strip().upper(): col for col in df.columns}
        def get_col(target_list):
            for t in target_list:
                if t in col_map: return col_map[t]
            return None
        
        po_col = get_col(["PO_NUMBER", "PO", "ORDER"])
        proj_col = get_col(["PROJECT", "PROJECT_ID"])
        comp_col = get_col(["COMPONENT_ID", "COMPONENT"])
        part_col = get_col(["PART_ID", "PART"])
        qty_col = get_col(["TOTAL_QTY", "QUANTITY", "QTY", "TARGET_QTY"])
        start_col = get_col(["START_DATE", "START"])
        due_col = get_col(["DUE_DATE", "END_DATE", "DUE"])
        
        rename_dict = {}
        if po_col: rename_dict[po_col] = 'po_number'
        if proj_col: rename_dict[proj_col] = 'project_id'
        if comp_col: rename_dict[comp_col] = 'component_id'
        if part_col: rename_dict[part_col] = 'part_id'
        if qty_col: rename_dict[qty_col] = 'total_qty'
        if start_col: rename_dict[start_col] = 'start_date'
        if due_col: rename_dict[due_col] = 'due_date'
        
        df.rename(columns=rename_dict, inplace=True)
        if 'start_date' in df.columns: df['start_date'] = pd.to_datetime(df['start_date'], errors='coerce')
        if 'due_date' in df.columns: df['due_date'] = pd.to_datetime(df['due_date'], errors='coerce')
        if 'total_qty' in df.columns: df['total_qty'] = pd.to_numeric(df['total_qty'], errors='coerce').fillna(0)
        
        return df
    except Exception:
        return pd.DataFrame()


def collapse_po_record(po_df):
    """Collapse multiple logistics rows for the same PO into a single dict.

    For multi-part POs the logistics file may contain one row per part.
    This aggregates them: sum total_qty, earliest start_date, latest due_date,
    and takes the first value for all other columns.
    """
    if po_df.empty:
        return {}
    if len(po_df) == 1:
        return po_df.iloc[0].to_dict()

    row = po_df.iloc[0].to_dict()
    if 'total_qty' in po_df.columns:
        row['total_qty'] = po_df['total_qty'].sum()
    if 'start_date' in po_df.columns:
        valid = po_df['start_date'].dropna()
        if not valid.empty:
            row['start_date'] = valid.min()
    if 'due_date' in po_df.columns:
        valid = po_df['due_date'].dropna()
        if not valid.empty:
            row['due_date'] = valid.max()
    return row


def evaluate_po_step_milestones(po_rec_df, current_cum, avg_daily_rate, last_actual_date, today):
    """Evaluate a multi-part PO against its chronological milestone deadlines (step-function).

    When one po_number has multiple rows in the logistics plan (different parts with
    different due dates), collapsing to max(due_date) + sum(qty) misses intermediate
    deadlines. This function iterates each milestone chronologically and flags the
    first one the current trajectory will miss.

    Returns a dict with status, first_failing_due_date, first_failing_part, and
    projected/required quantities at the failing milestone — or None if insufficient data.
    """
    if po_rec_df.empty or len(po_rec_df) < 2:
        return None  # Single-row PO: no intermediate milestones to evaluate
    if avg_daily_rate <= 0 or last_actual_date is None:
        return None

    required_cols = {'due_date', 'total_qty'}
    if not required_cols.issubset(po_rec_df.columns):
        return None

    work = po_rec_df[['due_date', 'total_qty'] + (['part_id'] if 'part_id' in po_rec_df.columns else [])].copy()
    work['due_date'] = pd.to_datetime(work['due_date'], errors='coerce')
    work['total_qty'] = pd.to_numeric(work['total_qty'], errors='coerce').fillna(0)
    work = work.dropna(subset=['due_date'])

    # Group in case multiple parts share an exact due date
    group_cols = ['due_date']
    agg_dict = {'total_qty': 'sum'}
    if 'part_id' in work.columns:
        agg_dict['part_id'] = 'first'
    milestones = work.groupby('due_date', as_index=False).agg(agg_dict).sort_values('due_date')
    milestones['cumulative_target'] = milestones['total_qty'].cumsum()

    last_date = pd.Timestamp(last_actual_date)

    for _, ms in milestones.iterrows():
        days_to_ms = (ms['due_date'] - last_date).days
        projected = current_cum + (avg_daily_rate * max(days_to_ms, 0))
        if projected < ms['cumulative_target']:
            due_str = ms['due_date'].strftime('%Y-%m-%d')
            ms_status = "Late" if ms['due_date'].date() < today else "At Risk"
            part_label = str(ms.get('part_id', '')) if 'part_id' in ms.index else ''
            return {
                'status': ms_status,
                'first_failing_due_date': due_str,
                'first_failing_part': part_label,
                'projected_at_milestone': round(projected, 0),
                'required_at_milestone': round(float(ms['cumulative_target']), 0),
            }

    return {'status': 'On Track', 'first_failing_due_date': None, 'first_failing_part': None}


def load_all_data_cr(files):
    """
    Loads and standardises production shot data.
    Column aliases are identical to run_rate_utils.load_all_data so that
    the same import files (DB names or internal names) work in both apps.

    DB primary names (from database export):
        EQUIPMENT_CODE  → tool_id
        CT              → actual_ct
        APPROVED_CT     → approved_ct
        LOCAL_SHOT_TIME → shot_time   (DD/MM/YYYY HH:MM:SS.fff, dayfirst)
        SUPPLIER_ID     → supplier_id
        PLANT_ID        → plant_id
        MATERIAL        → material
        PART_ID         → part_id
        PART_NAME       → part_name
        TOOLING_TYPE    → tooling_type
        PROJECT_ID      → project_id
        PO_NUMBER       → po_number
        WORKING_CAVITIES→ working_cavities
    """
    df_list = []
    for file in files:
        try:
            df = (pd.read_excel(file)
                  if file.name.endswith(('.xls', '.xlsx'))
                  else pd.read_csv(file))

            col_map = {col.strip().upper(): col for col in df.columns}

            def get_col(*targets):
                for t in targets:
                    found = col_map.get(t.strip().upper())
                    if found is not None:
                        return found
                return None

            # ── Tool ID ────────────────────────────────────────────────────────
            tool_col = get_col("EQUIPMENT_CODE", "TOOLING ID", "EQUIPMENT CODE",
                               "TOOL_ID", "TOOL", "tool_id")
            if tool_col:
                df.rename(columns={tool_col: "tool_id"}, inplace=True)

            # ── Actual CT ──────────────────────────────────────────────────────
            act_ct_col = get_col("CT", "ACTUAL CT", "ACTUAL_CT",
                                 "CYCLE TIME", "actual_ct")
            if act_ct_col:
                df.rename(columns={act_ct_col: "actual_ct"}, inplace=True)

            # ── Approved CT ────────────────────────────────────────────────────
            app_ct_col = get_col("APPROVED_CT", "APPROVED CT", "TARGET CT",
                                 "TARGET_CT", "STANDARD CT", "STD CT", "approved_ct")
            if app_ct_col:
                df.rename(columns={app_ct_col: "approved_ct"}, inplace=True)

            # ── Shot Timestamp ─────────────────────────────────────────────────
            # LOCAL_SHOT_TIME is DD/MM/YYYY (dayfirst=True); all other aliases
            # use YYYY-MM-DD or auto-inferred format (dayfirst=False).
            if {"YEAR", "MONTH", "DAY", "TIME"}.issubset(set(col_map.keys())):
                datetime_str = (
                    df[col_map["YEAR"]].astype(str) + "-"
                    + df[col_map["MONTH"]].astype(str) + "-"
                    + df[col_map["DAY"]].astype(str) + " "
                    + df[col_map["TIME"]].astype(str)
                )
                df["shot_time"] = pd.to_datetime(datetime_str, errors="coerce")
            else:
                time_col = get_col("LOCAL_SHOT_TIME", "SHOT_TIME", "SHOT TIME",
                                   "TIMESTAMP", "DATE", "TIME", "shot_time")
                if time_col:
                    is_db_col = (time_col.strip().upper() == "LOCAL_SHOT_TIME")
                    parsed = pd.to_datetime(
                        df[time_col], dayfirst=is_db_col, errors="coerce"
                    )
                    if parsed.isna().mean() > 0.5:
                        parsed = pd.to_datetime(
                            df[time_col], format="mixed",
                            dayfirst=is_db_col, errors="coerce"
                        )
                    df["shot_time"] = parsed

            # ── Hierarchy / filter columns ─────────────────────────────────────
            sup_col = get_col("SUPPLIER_ID", "SUPPLIER_NAME", "SUPPLIER NAME",
                              "SUPPLIER", "supplier_id")
            if sup_col and sup_col != "supplier_id":
                df.rename(columns={sup_col: "supplier_id"}, inplace=True)

            plt_col = get_col("PLANT_ID", "PLANT", "FACTORY", "plant_id")
            if plt_col and plt_col != "plant_id":
                df.rename(columns={plt_col: "plant_id"}, inplace=True)

            mat_col = get_col("MATERIAL", "MAT", "RESIN", "material")
            if mat_col and mat_col != "material":
                df.rename(columns={mat_col: "material"}, inplace=True)

            part_id_col = get_col("PART_ID", "PART", "part_id")
            if part_id_col and part_id_col != "part_id":
                df.rename(columns={part_id_col: "part_id"}, inplace=True)

            part_name_col = get_col("PART_NAME", "PART NAME", "part_name")
            if part_name_col and part_name_col != "part_name":
                df.rename(columns={part_name_col: "part_name"}, inplace=True)

            tt_col = get_col("TOOLING_TYPE", "TOOLING TYPE", "tooling_type")
            if tt_col and tt_col != "tooling_type":
                df.rename(columns={tt_col: "tooling_type"}, inplace=True)

            proj_col = get_col("PROJECT_ID", "PROJECT", "PROJECT NAME",
                               "PROJECT_NAME", "project_id")
            if proj_col and proj_col != "project_id":
                df.rename(columns={proj_col: "project_id"}, inplace=True)

            po_col = get_col("PO_NUMBER", "PO", "ORDER")
            if po_col:
                df.rename(columns={po_col: "po_number"}, inplace=True)

            cav_col = get_col("WORKING_CAVITIES", "WORKING CAVITIES", "CAVITIES")
            if cav_col:
                df.rename(columns={cav_col: "working_cavities"}, inplace=True)

            area_col = get_col("PLANT_AREA", "PLANT AREA", "AREA")
            if area_col:
                df.rename(columns={area_col: "plant_area"}, inplace=True)

            if "shot_time" in df.columns and "actual_ct" in df.columns:
                df["shot_time"] = pd.to_datetime(df["shot_time"], errors="coerce")
                df["actual_ct"] = pd.to_numeric(df["actual_ct"], errors="coerce")
                df.dropna(subset=["shot_time", "actual_ct"], inplace=True)
                df_list.append(df)

        except Exception:
            continue

    if not df_list:
        return pd.DataFrame()

    df_final = pd.concat(df_list, ignore_index=True)

    if 'tool_id' not in df_final.columns:
        df_final['tool_id'] = 'Unknown'
    df_final['tool_id'] = df_final['tool_id'].fillna('Unknown').astype(str)

    # Normalise all hierarchy columns — strip whitespace, null→Unknown.
    # Identical to run_rate_utils so the same files work in both apps.
    for _col in ['supplier_id', 'plant_id', 'project_id', 'material',
                 'part_id', 'tooling_type', 'po_number']:
        if _col in df_final.columns:
            df_final[_col] = (df_final[_col].astype(str).str.strip()
                              .replace({'nan': 'Unknown', 'none': 'Unknown',
                                        'None': 'Unknown', 'NaN': 'Unknown',
                                        'NAT': 'Unknown', '': 'Unknown'}))

    return df_final


# ==============================================================================
# --- CORE CALCULATION ENGINE ---
# ==============================================================================

class CapacityRiskCalculator:
    def __init__(self, df: pd.DataFrame, tolerance: float, downtime_gap_tolerance: float, 
                 run_interval_hours: float, target_output_perc: float = 100.0, 
                 default_cavities: int = 1, remove_maintenance: bool = False, **kwargs):
        
        self.df_raw = df.copy()
        self.tolerance = tolerance
        self.downtime_gap_tolerance = downtime_gap_tolerance
        self.run_interval_hours = run_interval_hours
        self.target_output_perc = target_output_perc
        self.default_cavities = default_cavities
        self.remove_maintenance = remove_maintenance
        self.results = self._calculate_metrics()

    def _calculate_metrics(self) -> dict:
        df = self.df_raw.copy()
        if df.empty: return {}

        if self.remove_maintenance and 'plant_area' in df.columns:
            df = df[~df['plant_area'].astype(str).str.lower().isin(['maintenance', 'warehouse'])].copy()
            if df.empty: return {}

        if 'approved_ct' not in df.columns: df['approved_ct'] = df['actual_ct'].median() 
        if 'working_cavities' not in df.columns: df['working_cavities'] = self.default_cavities
        
        df['approved_ct'] = pd.to_numeric(df['approved_ct'], errors='coerce').fillna(1)
        df['working_cavities'] = pd.to_numeric(df['working_cavities'], errors='coerce').fillna(self.default_cavities)
        
        # Ensure positive Approved CT
        df.loc[df['approved_ct'] <= 0, 'approved_ct'] = 1
        
        # Sort fundamentally respects tool_id first so rolled-up logic functions perfectly
        df = df.sort_values(["tool_id", "shot_time"]).reset_index(drop=True)

        # 1. Run Identification (Grouped by tool_id safely)
        df['time_diff_sec'] = df.groupby('tool_id')['shot_time'].diff().dt.total_seconds().fillna(0)
        
        # Initialize the first shot per tool to actual_ct
        mask_first_shot = df['tool_id'] != df['tool_id'].shift(1)
        df.loc[mask_first_shot, 'time_diff_sec'] = df.loc[mask_first_shot, 'actual_ct']

        is_new_run = df['time_diff_sec'] > (self.run_interval_hours * 3600)
        
        # Generates globally unique Run IDs securely isolated by Tool boundaries
        df['run_id'] = (is_new_run | mask_first_shot).cumsum()

        # 2. Mode CT & Limits
        run_modes = df[df['actual_ct'] < 999.9].groupby('run_id')['actual_ct'].apply(
            _get_stable_mode
        )
        df['mode_ct'] = df['run_id'].map(run_modes)
        lower_limit = df['mode_ct'] * (1 - self.tolerance)
        upper_limit = df['mode_ct'] * (1 + self.tolerance)
        
        df['mode_lower'] = lower_limit
        df['mode_upper'] = upper_limit

        # 3. Approved CT
        run_approved_cts = df.groupby('run_id')['approved_ct'].apply(
            _get_stable_mode
        )
        df['approved_ct_for_run'] = df['run_id'].map(run_approved_cts)
        
        # 4. Stop Detection (Isolated explicitly by tool limits)
        df['next_shot_time_diff'] = df.groupby('tool_id')['time_diff_sec'].shift(-1).fillna(0)
        
        is_time_gap = df['next_shot_time_diff'] > (df['actual_ct'] + self.downtime_gap_tolerance)
        is_abnormal = ((df['actual_ct'] < lower_limit) | (df['actual_ct'] > upper_limit))
        is_hard_stop = df['actual_ct'] >= 999.9

        df['stop_flag'] = np.where(is_time_gap | is_abnormal | is_hard_stop, 1, 0)

        # Reset stop_flag for first shots and new-run shots, but only when the
        # CT is within 5× mode — a machine idle for days is genuine downtime
        # and must remain flagged regardless of run boundary. (Aligned with RR.)
        startup_ct_ok = df['actual_ct'] < (df['mode_ct'] * 5)
        df.loc[(mask_first_shot | is_new_run) & startup_ct_ok, 'stop_flag'] = 0

        # Protect against stop events bleeding over tool boundaries
        df['prev_stop_flag'] = df.groupby('tool_id')['stop_flag'].shift(1, fill_value=0)
        df['stop_event'] = (df["stop_flag"] == 1) & (df["prev_stop_flag"] == 0)

        df['adj_ct_sec'] = df['actual_ct']
        df.loc[is_time_gap, 'adj_ct_sec'] = df['next_shot_time_diff']
        
        # --- Metrics Calculation ---
        run_durations = []
        run_opt_parts = []
        
        for _, run_df in df.groupby('run_id'):
            if not run_df.empty:
                start = run_df['shot_time'].min()
                end = run_df['shot_time'].max()
                last_ct = run_df.iloc[-1]['actual_ct']
                duration = (end - start).total_seconds() + last_ct
                run_durations.append(duration)
                
                # Calculate Optimal Output precisely per run to prevent multi-tool averages from skewing output
                r_ct = run_df['approved_ct_for_run'].iloc[0]
                r_cav = run_df['working_cavities'].max()
                run_opt_parts.append((duration / r_ct) * r_cav)
        
        total_runtime_sec = sum(run_durations)
        optimal_output_parts = sum(run_opt_parts)

        prod_df = df[df['stop_flag'] == 0].copy()
        production_time_sec = prod_df['actual_ct'].sum()
        downtime_sec = max(0, total_runtime_sec - production_time_sec)

        stops = df['stop_event'].sum()
        mttr_min = (downtime_sec / 60 / stops) if stops > 0 else 0
        stability_index = (production_time_sec / total_runtime_sec * 100) if total_runtime_sec > 0 else 100.0

        # --- Capacity Logic ---
        actual_output_parts = prod_df['working_cavities'].sum()
        target_output_parts = optimal_output_parts * (self.target_output_perc / 100.0)

        true_loss_parts = optimal_output_parts - actual_output_parts
        
        # Initialize variables
        capacity_gain_fast_parts = 0.0
        capacity_loss_slow_parts = 0.0
        capacity_gain_fast_sec = 0.0
        capacity_loss_slow_sec = 0.0

        if not prod_df.empty:
            # Inefficiency Calculation
            prod_df['parts_delta'] = ((prod_df['approved_ct_for_run'] - prod_df['actual_ct']) / prod_df['approved_ct_for_run']) * prod_df['working_cavities']
            
            capacity_gain_fast_parts = prod_df.loc[prod_df['parts_delta'] > 0, 'parts_delta'].sum()
            capacity_loss_slow_parts = abs(prod_df.loc[prod_df['parts_delta'] < 0, 'parts_delta'].sum())
            
            # Inefficiency Time Calculation
            prod_df['time_delta'] = prod_df['approved_ct_for_run'] - prod_df['actual_ct']
            capacity_gain_fast_sec = prod_df.loc[prod_df['time_delta'] > 0, 'time_delta'].sum()
            capacity_loss_slow_sec = abs(prod_df.loc[prod_df['time_delta'] < 0, 'time_delta'].sum())

        net_cycle_loss_parts = capacity_loss_slow_parts - capacity_gain_fast_parts
        capacity_loss_downtime_parts = true_loss_parts - net_cycle_loss_parts
        
        net_cycle_loss_sec = capacity_loss_slow_sec - capacity_gain_fast_sec
        
        # Enhanced to use time-exact loss instead of multiplying by average CT
        total_capacity_loss_sec = downtime_sec + capacity_loss_slow_sec
        
        gap_to_target_parts = actual_output_parts - target_output_parts
        capacity_loss_vs_target_parts = max(0, -gap_to_target_parts)
        
        total_shots = len(df)
        stop_count_shots = df['stop_flag'].sum()
        normal_shots = total_shots - stop_count_shots
        
        run_rate_efficiency = (normal_shots / total_shots * 100) if total_shots > 0 else 0
        capacity_efficiency = (actual_output_parts / optimal_output_parts) if optimal_output_parts > 0 else 0

        # Shot Typing
        epsilon = 0.001
        conditions = [
            df['stop_flag'] == 1,
            df['actual_ct'] > (df['approved_ct_for_run'] + epsilon), 
            df['actual_ct'] < (df['approved_ct_for_run'] - epsilon)
        ]
        choices = ['Downtime (Stop)', 'Slow Cycle', 'Fast Cycle']
        df['shot_type'] = np.select(conditions, choices, default='On Target')

        return {
            "processed_df": df,
            "total_runtime_sec": total_runtime_sec,
            "production_time_sec": production_time_sec,
            "downtime_sec": downtime_sec,
            "mttr_min": mttr_min,
            "stability_index": stability_index,
            "stops": stops,
            "optimal_output_parts": optimal_output_parts,
            "actual_output_parts": actual_output_parts,
            "target_output_parts": target_output_parts,
            "capacity_loss_downtime_parts": capacity_loss_downtime_parts,
            "capacity_loss_slow_parts": capacity_loss_slow_parts,
            "capacity_gain_fast_parts": capacity_gain_fast_parts,
            "total_capacity_loss_parts": true_loss_parts,
            "total_capacity_loss_sec": total_capacity_loss_sec,
            "gap_to_target_parts": gap_to_target_parts,
            "capacity_loss_vs_target_parts": capacity_loss_vs_target_parts,
            "efficiency_rate": run_rate_efficiency,
            "capacity_efficiency": capacity_efficiency,
            "total_shots": total_shots,
            "normal_shots": normal_shots,
            "mtbf_min": (production_time_sec / 60 / stops) if stops > 0 else (production_time_sec / 60)
        }

def calculate_run_summaries(df_period, config):
    """
    Calculates per-run metrics from a pre-processed dataframe (stop_flag already set).
    Does NOT re-run CapacityRiskCalculator — reads pre-computed columns directly,
    eliminating boundary/slice discrepancies between tabs.
    """
    summary_list = []
    if 'run_id' not in df_period.columns:
        return pd.DataFrame()

    # Safety fallback — if somehow called on unprocessed data
    if 'stop_flag' not in df_period.columns:
        calc = CapacityRiskCalculator(df_period, **config)
        df_period = calc.results.get('processed_df', df_period)

    for r_id, df_run in df_period.groupby('run_id'):
        if df_run.empty:
            continue

        start    = df_run['shot_time'].min()
        end      = df_run['shot_time'].max()
        last_ct  = float(df_run.iloc[-1]['actual_ct'])
        duration = (end - start).total_seconds() + last_ct

        r_ct  = (df_run['approved_ct_for_run'].iloc[0]
                 if 'approved_ct_for_run' in df_run.columns
                 else df_run['approved_ct'].iloc[0])
        r_cav = df_run['working_cavities'].max() if 'working_cavities' in df_run.columns else 1
        opt_parts = (duration / r_ct) * r_cav if r_ct > 0 else 0

        prod_df   = df_run[df_run['stop_flag'] == 0]
        prod_time = float(prod_df['actual_ct'].sum())
        downtime  = max(0, duration - prod_time)
        act_output = float(prod_df['working_cavities'].sum()) if 'working_cavities' in prod_df.columns else float(len(prod_df))
        tgt_output = opt_parts * (config.get('target_output_perc', 100.0) / 100.0)

        cap_gain_fast = cap_loss_slow = 0.0
        if not prod_df.empty and 'approved_ct_for_run' in prod_df.columns:
            parts_delta = (
                (prod_df['approved_ct_for_run'] - prod_df['actual_ct'])
                / prod_df['approved_ct_for_run'].replace(0, np.nan)
            ) * (prod_df['working_cavities'] if 'working_cavities' in prod_df.columns else 1)
            cap_gain_fast = float(parts_delta[parts_delta > 0].sum())
            cap_loss_slow = float(abs(parts_delta[parts_delta < 0].sum()))

        true_loss    = opt_parts - act_output
        net_cycle    = cap_loss_slow - cap_gain_fast
        loss_dt      = true_loss - net_cycle
        stops        = int(df_run['stop_event'].sum()) if 'stop_event' in df_run.columns else 0
        total_shots  = len(df_run)
        normal_shots = len(prod_df)

        summary_list.append({
            'run_id':                      r_id,
            'tool_ids':                    (', '.join(df_run['tool_id'].astype(str).unique())
                                            if 'tool_id' in df_run.columns else 'Unknown'),
            'start_time':                  start,
            'end_time':                    end,
            'total_shots':                 total_shots,
            'normal_shots':                normal_shots,
            'stop_events':                 stops,
            'stopped_shots':               total_shots - normal_shots,
            'mode_ct':                     float(df_run['mode_ct'].iloc[0]) if 'mode_ct' in df_run.columns else 0.0,
            'mode_lower':                  float(df_run['mode_lower'].iloc[0]) if 'mode_lower' in df_run.columns else 0.0,
            'mode_upper':                  float(df_run['mode_upper'].iloc[0]) if 'mode_upper' in df_run.columns else 0.0,
            'total_runtime_sec':           duration,
            'production_time_sec':         prod_time,
            'downtime_sec':                downtime,
            'total_capacity_loss_sec':     downtime + cap_loss_slow,
            'optimal_output_parts':        opt_parts,
            'target_output_parts':         tgt_output,
            'actual_output_parts':         act_output,
            'capacity_loss_downtime_parts': loss_dt,
            'capacity_loss_slow_parts':    cap_loss_slow,
            'capacity_gain_fast_parts':    cap_gain_fast,
            'total_capacity_loss_parts':   true_loss,
            'mttr_min':                    (downtime / 60 / stops) if stops > 0 else 0.0,
            'stability_index':             (prod_time / duration * 100) if duration > 0 else 100.0,
        })

    if not summary_list:
        return pd.DataFrame()
    df_summary = pd.DataFrame(summary_list).sort_values('start_time').reset_index(drop=True)
    df_summary['display_run_id'] = range(1, len(df_summary) + 1)
    return df_summary

# ==============================================================================
# --- AGGREGATION, PREDICTION & RISK LOGIC ---
# ==============================================================================

def get_aggregated_data(df, freq_mode, config):
    """
    Aggregates a pre-processed dataframe by period.
    Calls calculate_run_summaries (which reads pre-computed columns) — does NOT
    re-instantiate CapacityRiskCalculator per period, avoiding metric drift.
    """
    if df.empty or 'shot_time' not in df.columns:
        return pd.DataFrame()

    if freq_mode == 'Daily':
        df = df.copy()
        df['Period_Lbl'] = df['shot_time'].dt.date.astype(str)
    elif freq_mode == 'Weekly':
        df = df.copy()
        df['Period_Lbl'] = df['shot_time'].dt.to_period('W').astype(str)
    elif freq_mode == 'Monthly':
        df = df.copy()
        df['Period_Lbl'] = df['shot_time'].dt.to_period('M').astype(str)
    elif freq_mode == 'Hourly':
        df = df.copy()
        df['Period_Lbl'] = df['shot_time'].dt.floor('h').astype(str)
    elif freq_mode == 'by Run':
        df = df.copy()
        df['Period_Lbl'] = df['run_id'].astype(str)
    else:
        return pd.DataFrame()

    agg_rows = []
    for period, df_period in df.groupby('Period_Lbl'):
        run_breakdown_df = calculate_run_summaries(df_period, config)
        if run_breakdown_df.empty:
            continue

        total_runtime = run_breakdown_df['total_runtime_sec'].sum()
        prod_time     = run_breakdown_df['production_time_sec'].sum()
        downtime      = run_breakdown_df['downtime_sec'].sum()
        opt_output    = run_breakdown_df['optimal_output_parts'].sum()
        tgt_output    = (run_breakdown_df['target_output_parts'].sum()
                         if 'target_output_parts' in run_breakdown_df.columns
                         else opt_output * (config.get('target_output_perc', 100.0) / 100.0))
        act_output    = run_breakdown_df['actual_output_parts'].sum()
        loss_dt       = run_breakdown_df['capacity_loss_downtime_parts'].sum()
        loss_slow     = run_breakdown_df['capacity_loss_slow_parts'].sum()
        gain_fast     = run_breakdown_df['capacity_gain_fast_parts'].sum()
        total_loss    = run_breakdown_df['total_capacity_loss_parts'].sum()
        total_shots   = run_breakdown_df['total_shots'].sum()
        normal_shots  = run_breakdown_df['normal_shots'].sum()

        agg_rows.append({
            'Period':             period,
            'Runs':               len(run_breakdown_df),
            'Actual Output':      act_output,
            'Optimal Output':     opt_output,
            'Target Output':      tgt_output,
            'Downtime Loss':      loss_dt,
            'Slow Loss':          loss_slow,
            'Fast Gain':          gain_fast,
            'Net Cycle Loss':     loss_slow - gain_fast,
            'Total Loss':         total_loss,
            'Gap to Target':      max(0, tgt_output - act_output),
            'Run Time':           format_seconds_to_dhm(total_runtime),
            'Downtime':           format_seconds_to_dhm(downtime),
            'Run Time Sec':       total_runtime,
            'Production Time Sec': prod_time,
            'Downtime Sec':       downtime,
            'Total Shots':        total_shots,
            'Normal Shots':       normal_shots,
            'Downtime Shots':     total_shots - normal_shots,
        })

    if not agg_rows:
        return pd.DataFrame()
    df_agg = pd.DataFrame(agg_rows).sort_values('Period').reset_index(drop=True)
    if freq_mode == 'by Run':
        df_agg['Period'] = [f"Run {i}" for i in range(1, len(df_agg) + 1)]
    return df_agg

def generate_po_periodic_data(df_bar_view, po_record, freq_mode, config, working_days_per_week, working_hours_per_day):
    """Generates periodic aggregated data spanning the full PO timeline."""
    start_date = po_record.get('start_date')
    due_date = po_record.get('due_date')
    if pd.isna(start_date) or pd.isna(due_date): return pd.DataFrame()
    
    start_date = pd.to_datetime(start_date)
    due_date = pd.to_datetime(due_date)
    total_qty = po_record.get('total_qty', 0)
    
    # Calculate uniform demand spread
    total_calendar_days = (due_date - start_date).days
    if total_calendar_days <= 0: total_calendar_days = 1
    
    total_weeks = total_calendar_days / 7.0
    total_working_days = total_weeks * working_days_per_week
    
    daily_demand = total_qty / total_working_days if total_working_days > 0 else 0
    weekly_demand = daily_demand * working_days_per_week
    monthly_demand = weekly_demand * 4.33
    
    # Calculate Configured Max Capacity based on optimal cycle times
    avg_ct = df_bar_view['approved_ct'].mean() if (not df_bar_view.empty and 'approved_ct' in df_bar_view.columns) else 1
    if pd.isna(avg_ct) or avg_ct <= 0: avg_ct = 1
    cav = df_bar_view['working_cavities'].max() if (not df_bar_view.empty and 'working_cavities' in df_bar_view.columns) else 1
    
    hourly_cap = (3600 / avg_ct) * cav
    
    # Process Actuals Data
    agg_df = get_aggregated_data(df_bar_view, freq_mode, config)
    
    # Calculate timeline bounds to match the Burn-Up chart scope (Start -> Due or Late Finish)
    current_cum = agg_df['Actual Output'].sum() if not agg_df.empty else 0
    last_actual_date = pd.to_datetime(df_bar_view['shot_time'].max()).date() if not df_bar_view.empty else start_date.date()
    
    days_elapsed = (last_actual_date - start_date.date()).days + 1
    if days_elapsed <= 0: days_elapsed = 1
    avg_daily_rate = current_cum / days_elapsed
    
    remaining_qty = total_qty - current_cum
    max_proj_days = 0
    if remaining_qty > 0 and avg_daily_rate > 0:
        max_proj_days = int(remaining_qty / avg_daily_rate) + 1
        max_proj_days = min(max_proj_days, 365) # cap prediction limits
        
    projected_end_date = last_actual_date + timedelta(days=max_proj_days)
    end_timeline_date = max(due_date.date(), projected_end_date)
    
    # Generate continuous empty timeline
    if freq_mode == 'Daily':
        full_periods = pd.date_range(start=start_date.date(), end=end_timeline_date, freq='D').date
        df_full = pd.DataFrame({'Period': full_periods})
        df_full['Estimated Demand'] = daily_demand
        df_full['Configured Max Capacity'] = hourly_cap * working_hours_per_day
        if not agg_df.empty: agg_df['Period'] = pd.to_datetime(agg_df['Period']).dt.date
    elif freq_mode == 'Weekly':
        full_periods = pd.period_range(start=start_date, end=end_timeline_date, freq='W').astype(str)
        df_full = pd.DataFrame({'Period': full_periods})
        df_full['Estimated Demand'] = weekly_demand
        df_full['Configured Max Capacity'] = hourly_cap * working_hours_per_day * working_days_per_week
    elif freq_mode == 'Monthly':
        full_periods = pd.period_range(start=start_date, end=end_timeline_date, freq='M').astype(str)
        df_full = pd.DataFrame({'Period': full_periods})
        df_full['Estimated Demand'] = monthly_demand
        df_full['Configured Max Capacity'] = hourly_cap * working_hours_per_day * working_days_per_week * 4.33
    else:
        df_full = pd.DataFrame()

    # Merge full timeline with available actuals, filling gaps with zero
    if not df_full.empty:
        df_full['Period'] = df_full['Period'].astype(str)
        if not agg_df.empty:
            agg_df['Period'] = agg_df['Period'].astype(str)
            final_df = pd.merge(df_full, agg_df[['Period', 'Actual Output']], on='Period', how='left')
            final_df['Actual Output'] = final_df['Actual Output'].fillna(0)
        else:
            final_df = df_full
            final_df['Actual Output'] = 0
        return final_df
        
    return agg_df

def generate_po_prediction_data(df_po_shots, po_record, config, working_days=5, working_hours=24):
    """Generates time-series data specifically for PO Burn-up charting."""
    if pd.isna(po_record.get('start_date')) or pd.isna(po_record.get('due_date')):
        return None
        
    start_date = po_record['start_date'].date() if isinstance(po_record['start_date'], pd.Timestamp) else pd.to_datetime(po_record['start_date']).date()
    due_date = po_record['due_date'].date() if isinstance(po_record['due_date'], pd.Timestamp) else pd.to_datetime(po_record['due_date']).date()
    total_qty = po_record.get('total_qty', 0)

    # Target Burnup Line (Ideal linear progress)
    total_days = (due_date - start_date).days
    if total_days <= 0: total_days = 1
    
    target_dates = [start_date + timedelta(days=i) for i in range(total_days + 1)]
    target_vals = [(total_qty / total_days) * i for i in range(total_days + 1)]

    # Get daily aggregations for actuals
    agg_daily = get_aggregated_data(df_po_shots, 'Daily', config) if not df_po_shots.empty else pd.DataFrame()
    
    if agg_daily.empty:
        return {
            'target_dates': target_dates, 'target_vals': target_vals,
            'actual_dates': pd.Series(dtype=object), 'actual_cum': pd.Series(dtype=float),
            'forecast_dates': [], 'forecast_avg': [], 'forecast_opt': [],
            'due_date': due_date, 'start_date': start_date, 'total_qty': total_qty,
            'current_cum': 0, 'avg_daily_rate': 0, 'opt_daily_rate': 0
        }
        
    agg_daily['Period'] = pd.to_datetime(agg_daily['Period']).dt.date
    agg_daily = agg_daily.sort_values('Period')
    agg_daily['Cumulative Actual'] = agg_daily['Actual Output'].cumsum()
    
    last_actual_date = agg_daily['Period'].max()
    current_cum = agg_daily['Cumulative Actual'].max()
    
    # Calculate optimal rate from the whole subset
    calc = CapacityRiskCalculator(df_po_shots, **config)
    res = calc.results
    
    days_elapsed = (last_actual_date - start_date).days + 1
    if days_elapsed <= 0: days_elapsed = 1
    
    avg_daily_rate = current_cum / days_elapsed
    opt_daily_rate = res['optimal_output_parts'] / days_elapsed if res else 0
    
    remaining_qty = total_qty - current_cum
    max_proj_days = 0
    
    if remaining_qty > 0:
        days_to_finish_avg = int(remaining_qty / avg_daily_rate) + 1 if avg_daily_rate > 0 else 30
        days_to_finish_opt = int(remaining_qty / opt_daily_rate) + 1 if opt_daily_rate > 0 else 30
            
        max_proj_days = max(days_to_finish_avg, days_to_finish_opt, (due_date - last_actual_date).days)
        max_proj_days = min(max_proj_days, 365) # cap projection at 1 year max
    else:
        max_proj_days = max(0, (due_date - last_actual_date).days)
    
    forecast_dates = [last_actual_date + timedelta(days=i) for i in range(max_proj_days + 1)]
    forecast_avg = [current_cum + (avg_daily_rate * i) for i in range(max_proj_days + 1)]
    forecast_opt = [current_cum + (opt_daily_rate * i) for i in range(max_proj_days + 1)]

    return {
        'target_dates': target_dates, 'target_vals': target_vals,
        'actual_dates': agg_daily['Period'], 'actual_cum': agg_daily['Cumulative Actual'],
        'forecast_dates': forecast_dates, 'forecast_avg': forecast_avg, 'forecast_opt': forecast_opt,
        'due_date': due_date, 'start_date': start_date, 'total_qty': total_qty,
        'current_cum': current_cum, 'avg_daily_rate': avg_daily_rate, 'opt_daily_rate': opt_daily_rate
    }

def generate_po_summary_board(df_po_shots, po_record):
    """Generates a structured summary DataFrame for a given PO."""
    if po_record is None or pd.isna(po_record.get('po_number')):
        return pd.DataFrame()
        
    # Gather tools involved uniquely
    involved_tools = ", ".join(sorted([str(t) for t in df_po_shots['tool_id'].unique()])) if (not df_po_shots.empty and 'tool_id' in df_po_shots.columns) else "Unknown"
    
    # Safely extract supplier and plant if they exist in the production mapping
    supplier = "Unknown"
    plant = "Unknown"
    if not df_po_shots.empty:
        if 'supplier_id' in df_po_shots.columns:
            sup_series = df_po_shots['supplier_id'].dropna()
            if not sup_series.empty: supplier = str(sup_series.iloc[0])
        if 'plant_id' in df_po_shots.columns:
            plt_series = df_po_shots['plant_id'].dropna()
            if not plt_series.empty: plant = str(plt_series.iloc[0])

    start_dt = pd.to_datetime(po_record.get('start_date')).strftime('%Y-%m-%d') if pd.notnull(po_record.get('start_date')) else 'N/A'
    due_dt = pd.to_datetime(po_record.get('due_date')).strftime('%Y-%m-%d') if pd.notnull(po_record.get('due_date')) else 'N/A'
    
    data = {
        "Purchase Order #": [po_record.get('po_number', 'N/A')],
        "Project": [po_record.get('project_id', 'N/A')],
        "Component": [po_record.get('component_id', 'N/A')],
        "Part": [po_record.get('part_id', 'N/A')],
        "Assigned Tooling(s)": [involved_tools],
        "Supplier": [supplier],
        "Plant": [plant],
        "Total Quantity": [po_record.get('total_qty', 0)],
        "Start Date": [start_dt],
        "Due Date": [due_dt]
    }
    
    return pd.DataFrame(data)

def generate_prediction_data(df_daily_agg, start_date, target_date, demand_target_total=None):
    """Fallback projection chart data generation."""
    if df_daily_agg.empty: return None

    df = df_daily_agg.copy()
    df['Period'] = pd.to_datetime(df['Period'])
    df = df.sort_values('Period')
    
    df['Cumulative Actual'] = df['Actual Output'].cumsum()
    
    last_historic_ts = df['Period'].max()
    last_historic_date = last_historic_ts.date() if hasattr(last_historic_ts, 'date') else last_historic_ts
    current_cumulative = df['Cumulative Actual'].max()
    
    days_with_data = (last_historic_date - df['Period'].min().date()).days + 1
    if days_with_data < 1: days_with_data = 1
    
    avg_daily_rate = df['Actual Output'].sum() / days_with_data
    peak_daily_rate = df['Actual Output'].quantile(0.90) if len(df) > 5 else df['Actual Output'].max()

    if isinstance(target_date, datetime):
        target_date = target_date.date()
        
    projection_days = (target_date - last_historic_date).days
    if projection_days < 1: projection_days = 0
    
    future_dates = [last_historic_date + timedelta(days=i) for i in range(projection_days + 1)]
    
    proj_avg = [current_cumulative + (avg_daily_rate * i) for i in range(len(future_dates))]
    proj_peak = [current_cumulative + (peak_daily_rate * i) for i in range(len(future_dates))]
    
    req_rate = 0
    proj_req = []
    if demand_target_total:
        remaining_qty = max(0, demand_target_total - current_cumulative)
        if projection_days > 0:
            req_rate = remaining_qty / projection_days
            proj_req = [current_cumulative + (req_rate * i) for i in range(len(future_dates))]
    
    return {
        'historic_dates': df['Period'],
        'historic_cum': df['Cumulative Actual'],
        'future_dates': future_dates,
        'proj_avg': proj_avg,
        'proj_peak': proj_peak,
        'proj_req': proj_req,
        'rates': {'avg': avg_daily_rate, 'peak': peak_daily_rate, 'req': req_rate}
    }

def calculate_capacity_risk_scores(df_all, config):
    risk_data = []
    for tool_id, df_tool in df_all.groupby('tool_id'):
        max_date = df_tool['shot_time'].max()
        cutoff_date = max_date - timedelta(weeks=4)
        df_period = df_tool[df_tool['shot_time'] >= cutoff_date].copy()
        
        if df_period.empty: continue
        
        calc = CapacityRiskCalculator(df_period, **config)
        res = calc.results
        if res['target_output_parts'] == 0: continue
        
        ach_perc = (res['actual_output_parts'] / res['target_output_parts']) * 100
        
        midpoint = cutoff_date + (max_date - cutoff_date) / 2
        df_late = df_period[df_period['shot_time'] >= midpoint]
        df_early = df_period[df_period['shot_time'] < midpoint]
        
        trend = "Stable"
        if not df_early.empty and not df_late.empty:
            c_early = CapacityRiskCalculator(df_early, **config).results
            c_late = CapacityRiskCalculator(df_late, **config).results
            early_rate = c_early['actual_output_parts'] / (c_early['total_runtime_sec']/3600) if c_early['total_runtime_sec'] > 0 else 0
            late_rate = c_late['actual_output_parts'] / (c_late['total_runtime_sec']/3600) if c_late['total_runtime_sec'] > 0 else 0
            
            if late_rate < early_rate * 0.95: trend = "Declining"
            elif late_rate > early_rate * 1.05: trend = "Improving"

        base_score = min(ach_perc, 100)
        if trend == "Declining": base_score -= 20
        
        risk_data.append({
            'Tool ID': tool_id,
            'Risk Score': max(0, base_score),
            'Achievement %': ach_perc,
            'Trend': trend,
            'Gap': res['gap_to_target_parts']
        })
    return pd.DataFrame(risk_data).sort_values('Risk Score')

# ==============================================================================
# --- NEW: AUTOMATED INSIGHTS & EXPORT ---
# ==============================================================================

def generate_capacity_insights(res, benchmark_mode):
    """Generates natural language summary of the capacity loss."""
    if not res: return {"overall": "No data available."}
    
    act = res['actual_output_parts']
    tgt = res['target_output_parts'] if benchmark_mode == "Target" else res['optimal_output_parts']
    diff = act - tgt
    
    status = "exceeded" if diff >= 0 else "missed"
    perc = abs(diff / tgt * 100) if tgt > 0 else 0
    
    color = "green" if diff >= 0 else "red"
    overall = f"Production <strong style='color:{color}'>{status}</strong> the {benchmark_mode} goal by <strong>{perc:.1f}%</strong> ({abs(diff):,.0f} parts)."
    
    # Driver Analysis
    drivers = []
    loss_dt = res['capacity_loss_downtime_parts']
    net_slow = res['capacity_loss_slow_parts'] - res['capacity_gain_fast_parts']
    
    total_loss_absolute = loss_dt + max(0, net_slow)
    
    if total_loss_absolute > 0:
        dt_share = (loss_dt / total_loss_absolute) * 100
        slow_share = (max(0, net_slow) / total_loss_absolute) * 100
        
        loss_term = "uncaptured capacity" if status == "exceeded" else "loss"
        driver_intro = "The primary constraint was" if status == "exceeded" else "The primary driver was"
        
        if dt_share > 60:
            drivers.append(f"{driver_intro} <strong>Downtime</strong>, accounting for <strong>{dt_share:.0f}%</strong> of the {loss_term}.")
            rec = "Focus on reducing Stop Events (MTBF) and improving Reaction Time (MTTR)."
        elif slow_share > 60:
            drivers.append(f"{driver_intro} <strong>Slow Cycle Time</strong>, accounting for <strong>{slow_share:.0f}%</strong> of the {loss_term}.")
            rec = "Investigate process parameters causing the machine to run slower than the Approved Cycle Time."
        else:
            drivers.append(f"Constraints were split between <strong>Downtime ({dt_share:.0f}%)</strong> and <strong>Slow Cycles ({slow_share:.0f}%)</strong>.")
            rec = "A balanced approach addressing both uptime and cycle speed is required."
    else:
        drivers.append("No significant capacity losses detected.")
        rec = "Maintain current performance standards."

    if res['capacity_gain_fast_parts'] > (res['actual_output_parts'] * 0.05):
        drivers.append(f"Note: Running faster than standard gained <strong>{res['capacity_gain_fast_parts']:,.0f}</strong> bonus parts.")

    return {"overall": overall, "drivers": " ".join(drivers), "recommendation": rec}

def generate_forecast_insights(pred_data, demand_target):
    """Generates insights for the forecast tab."""
    if not pred_data or demand_target <= 0:
        return "Please set a Demand Goal to generate a completion forecast."
    
    current_cum = pred_data['historic_cum'].iloc[-1]
    if current_cum >= demand_target:
        return f"🎉 <strong>Goal Achieved!</strong> Current output ({current_cum:,.0f}) has already exceeded the demand target ({demand_target:,.0f})."

    remaining = demand_target - current_cum
    rates = pred_data['rates']
    avg_rate = rates['avg']
    peak_rate = rates['peak']
    start_date = pred_data['future_dates'][0]

    def get_finish_date(rate):
        if rate <= 0: return None
        days = remaining / rate
        return start_date + timedelta(days=int(days))

    date_avg = get_finish_date(avg_rate)
    date_peak = get_finish_date(peak_rate)

    insight_html = f"""
    <ul style='margin-bottom:0;'>
        <li>To meet demand of <strong>{demand_target:,.0f}</strong>, you need <strong>{remaining:,.0f}</strong> more parts.</li>
    """

    if date_avg:
        insight_html += f"<li>At your <strong>Current Average Rate</strong> ({avg_rate:,.0f} parts/day), you will meet demand on <strong>{date_avg.strftime('%Y-%m-%d')}</strong>.</li>"
    else:
        insight_html += f"<li>At your <strong>Current Average Rate</strong>, you are not projected to meet demand (rate is 0 or negative).</li>"

    if date_peak:
        insight_html += f"<li>At <strong>Optimal/Peak Performance</strong> ({peak_rate:,.0f} parts/day), you could meet demand as early as <strong>{date_peak.strftime('%Y-%m-%d')}</strong>.</li>"
    
    insight_html += "</ul>"
    return insight_html

def prepare_and_generate_capacity_excel(df_view, config):
    """Generates the Excel export with formatted sheets."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        wb = writer.book
        
        fmt_header = wb.add_format({'bold':True,'bg_color':'#002060','font_color':'white','border':1})
        fmt_num = wb.add_format({'num_format':'#,##0','border':1})
        fmt_dec = wb.add_format({'num_format':'0.00','border':1})
        
        ws_sum = wb.add_worksheet("Management Summary")
        run_data = calculate_run_summaries(df_view, config)
        
        if not run_data.empty:
            ws_sum.write_row('A1', ['Start Time', 'End Time', 'Optimal', 'Actual', 'Loss (Downtime)', 'Loss (Slow)', 'Gain (Fast)'], fmt_header)
            
            for i, row in run_data.iterrows():
                ws_sum.write(i+1, 0, str(row['start_time']))
                ws_sum.write(i+1, 1, str(row['end_time']))
                ws_sum.write(i+1, 2, row['optimal_output_parts'], fmt_num)
                ws_sum.write(i+1, 3, row['actual_output_parts'], fmt_num)
                ws_sum.write(i+1, 4, row['capacity_loss_downtime_parts'], fmt_num)
                ws_sum.write(i+1, 5, row['capacity_loss_slow_parts'], fmt_num)
                ws_sum.write(i+1, 6, row['capacity_gain_fast_parts'], fmt_num)

        df_view.to_excel(writer, sheet_name="Raw Data", index=False)
        
    return output.getvalue()

def generate_mttr_mtbf_analysis(analysis_df):
    """Generates correlation text analysis for MTTR/MTBF drivers."""
    if analysis_df is None or analysis_df.empty or 'stop_events' not in analysis_df.columns:
        return "Not enough data to generate detailed correlation analysis."
        
    analysis_df_clean = analysis_df.dropna(subset=['stop_events', 'stability_index', 'mttr_min'])
    
    if analysis_df_clean.empty or len(analysis_df_clean) < 2:
        return "Insufficient data: At least 2 production runs with stop events are required to perform correlation analysis."

    period_col = 'display_run_id' if 'display_run_id' in analysis_df_clean.columns else 'run_id'
    
    df_calc = analysis_df_clean.rename(columns={
        'stop_events': 'stops', 
        'stability_index': 'stability', 
        'mttr_min': 'mttr',
        period_col: 'period'
    })
    
    stops_stability_corr = df_calc['stops'].corr(df_calc['stability'])
    mttr_stability_corr = df_calc['mttr'].corr(df_calc['stability'])
    
    corr_insight = ""
    primary_driver_is_frequency = False
    primary_driver_is_duration = False
    
    if not pd.isna(stops_stability_corr) and not pd.isna(mttr_stability_corr):
        if abs(stops_stability_corr) > abs(mttr_stability_corr) * 1.5:
            primary_driver = "the **frequency of stops**"
            primary_driver_is_frequency = True
        elif abs(mttr_stability_corr) > abs(stops_stability_corr) * 1.5:
            primary_driver = "the **duration of stops**"
            primary_driver_is_duration = True
        else:
            primary_driver = "both the **frequency and duration of stops**"
        corr_insight = (f"This analysis suggests that <strong>{primary_driver}</strong> has the strongest impact on overall stability.")
    
    example_insight = ""
    if primary_driver_is_frequency:
        highest_stops_row = df_calc.loc[df_calc['stops'].idxmax()]
        example_insight = (f"For example, Run {highest_stops_row['period']} recorded the most interruptions (<strong>{int(highest_stops_row['stops'])} stops</strong>). Prioritizing the root cause of these frequent events is recommended.")
    elif primary_driver_is_duration:
        highest_mttr_row = df_calc.loc[df_calc['mttr'].idxmax()]
        example_insight = (f"Run {highest_mttr_row['period']} experienced prolonged downtimes with an average repair time of <strong>{highest_mttr_row['mttr']:.1f} minutes</strong>. Investigating the cause of these prolonged stops is the top priority.")
    else:
        if not df_calc['mttr'].empty:
            highest_mttr_row = df_calc.loc[df_calc['mttr'].idxmax()]
            example_insight = (f"As an example, Run {highest_mttr_row['period']} experienced prolonged downtimes with an average repair time of <strong>{highest_mttr_row['mttr']:.1f} minutes</strong>, highlighting the impact of long stops.")
            
    return f"<div style='line-height: 1.6;'><p>{corr_insight}</p><p>{example_insight}</p></div>"

# ==============================================================================
# --- PLOTTING FUNCTIONS ---
# ==============================================================================

def plot_po_periodic_chart(agg_po, df_raw, bar_freq, track_mode):
    """Plots the periodic stacked bar chart for PO tracking vs Demand & Configured Capacity."""
    fig = go.Figure()
    
    # Determine what to stack the bars by based on what makes sense for the view mode
    breakdown_col = 'po_number' if 'Purchase Order' not in track_mode else 'tool_id'
    if breakdown_col not in df_raw.columns:
        breakdown_col = 'tool_id' # Safe fallback
        
    prod_df = df_raw[df_raw['stop_flag'] == 0].copy()
    
    if not prod_df.empty:
        # Assign accurate period bounds based on the frequency selected
        if bar_freq == 'Daily':
            prod_df['Period'] = prod_df['shot_time'].dt.date.astype(str)
        elif bar_freq == 'Weekly':
            prod_df['Period'] = prod_df['shot_time'].dt.to_period('W').astype(str)
        elif bar_freq == 'Monthly':
            prod_df['Period'] = prod_df['shot_time'].dt.to_period('M').astype(str)
        else:
            prod_df['Period'] = prod_df['shot_time'].dt.date.astype(str)
            
        bar_data = prod_df.groupby(['Period', breakdown_col])['working_cavities'].sum().reset_index()
        
        unique_segments = bar_data[breakdown_col].unique()
        colors = px.colors.qualitative.Pastel
        
        for i, segment in enumerate(unique_segments):
            seg_data = bar_data[bar_data[breakdown_col] == segment]
            fig.add_trace(go.Bar(
                x=seg_data['Period'], y=seg_data['working_cavities'],
                name=f'{segment}', marker_color=colors[i % len(colors)]
            ))
    else:
        # Safe fallback if raw data aggregation fails
        fig.add_trace(go.Bar(
            x=agg_po['Period'], y=agg_po['Actual Output'], 
            name='Actual Output', marker_color=PASTEL_COLORS['blue']
        ))
    
    # Overlay lines
    fig.add_trace(go.Scatter(
        x=agg_po['Period'], y=agg_po['Configured Max Capacity'], 
        name='Configured Max Capacity', mode='lines+markers', 
        line=dict(color=PASTEL_COLORS['green'], dash='dot', width=2)
    ))
    
    fig.add_trace(go.Scatter(
        x=agg_po['Period'], y=agg_po['Estimated Demand'], 
        name='Estimated PO Demand', mode='lines+markers', 
        line=dict(color=PASTEL_COLORS['red'], dash='dash', width=2)
    ))
    
    fig.update_layout(
        title=f"Periodic Production vs Demand ({bar_freq})",
        barmode='stack', hovermode="x unified", height=450,
        yaxis_title="Parts Output", xaxis_title="Period"
    )
    return fig

def plot_po_burnup(pred_data, po_logistics_df=None):
    """Plots the PO specific Burn-Up tracking chart with support for multiple PO Due Date annotations."""
    if not pred_data: return go.Figure()
    fig = go.Figure()
    
    # Target Burnup (Grey Dashed)
    if len(pred_data.get('target_dates', [])) > 0:
        fig.add_trace(go.Scatter(x=pred_data['target_dates'], y=pred_data['target_vals'], 
                                 mode='lines', name='Target Burn-up', line=dict(color='grey', dash='dash')))
                             
    # Actual Cumulative (Blue Line)
    if len(pred_data.get('actual_dates', [])) > 0:
        fig.add_trace(go.Scatter(x=pred_data['actual_dates'], y=pred_data['actual_cum'], 
                                 mode='lines+markers', name='Actual Accumulated', line=dict(color=PASTEL_COLORS['blue'], width=3)))
                             
    # Forecast Avg (Orange Dot)
    if pred_data.get('avg_daily_rate', 0) > 0 and len(pred_data.get('forecast_dates', [])) > 0:
        fig.add_trace(go.Scatter(x=pred_data['forecast_dates'], y=pred_data['forecast_avg'], 
                                 mode='lines', name=f"Forecast (Avg: {pred_data['avg_daily_rate']:.0f}/d)", line=dict(color=PASTEL_COLORS['orange'], dash='dot')))
                             
    # Forecast Opt (Green Dot)
    if pred_data.get('opt_daily_rate', 0) > 0 and len(pred_data.get('forecast_dates', [])) > 0:
        fig.add_trace(go.Scatter(x=pred_data['forecast_dates'], y=pred_data['forecast_opt'], 
                                 mode='lines', name=f"Forecast (Opt: {pred_data['opt_daily_rate']:.0f}/d)", line=dict(color=PASTEL_COLORS['green'], dash='dot')))
                             
    # Annotations - Loop through logistics records to print specific PO Due dates
    if po_logistics_df is not None and not po_logistics_df.empty:
        colors = px.colors.qualitative.Set1
        for idx, row in po_logistics_df.iterrows():
            if pd.notna(row['due_date']):
                due_ts = pd.to_datetime(row['due_date']).timestamp() * 1000
                po_num = row['po_number']
                fig.add_vline(x=due_ts, line_width=1.5, line_dash="dash", line_color=colors[idx % len(colors)], annotation_text=f"{po_num} Due")
    else:
        # Fallback to general due date
        if pd.notna(pred_data.get('due_date')):
            due_ts = pd.to_datetime(pred_data['due_date']).timestamp() * 1000
            fig.add_vline(x=due_ts, line_width=2, line_dash="dash", line_color="red", annotation_text="PO Due Date")
    
    fig.add_hline(y=pred_data.get('total_qty', 0), line_width=2, line_dash="solid", line_color="purple", annotation_text="Total Aggregated Qty")

    # ── Start date marker ─────────────────────────────────────────────────────
    start_dt = pd.to_datetime(pred_data.get('start_date', pd.Timestamp.now()))
    start_ts = start_dt.timestamp() * 1000
    fig.add_vline(x=start_ts, line_width=1.5, line_dash="dot", line_color="rgba(100,200,255,0.6)",
                  annotation_text="PO Start", annotation_position="top left",
                  annotation_font=dict(color="rgba(100,200,255,0.9)", size=11))

    # ── Due date — always visible with a filled annotation band ──────────────
    due_dt_pd = pred_data.get('due_date')
    if due_dt_pd is not None and pd.notna(due_dt_pd):
        due_ts = pd.to_datetime(due_dt_pd).timestamp() * 1000
        # Remove the basic vline already added above and re-add with stronger styling
        fig.add_vline(x=due_ts, line_width=2, line_dash="dash", line_color="rgba(231,76,60,0.9)",
                      annotation=dict(
                          text="<b>Due Date</b>",
                          font=dict(color="rgba(231,76,60,1.0)", size=12),
                          bgcolor="rgba(231,76,60,0.15)",
                          bordercolor="rgba(231,76,60,0.6)",
                          borderwidth=1,
                          borderpad=4,
                      ),
                      annotation_position="top right")

    # Force the X-axis bounds
    max_dt_target   = pd.to_datetime(pred_data['target_dates'][-1]) if len(pred_data.get('target_dates', [])) > 0 else start_dt
    max_dt_forecast = pd.to_datetime(pred_data['forecast_dates'][-1]) if len(pred_data.get('forecast_dates', [])) > 0 else max_dt_target
    end_dt = max(max_dt_target, max_dt_forecast)

    fig.update_layout(
        title="PO Target Burn-up vs Reality",
        hovermode="x unified",
        height=500,
        yaxis_title="Accumulated Parts Output",
        xaxis_title="Date",
        xaxis_range=[start_dt, end_dt],
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
    )
    return fig

def create_time_breakdown_donut(total_sec, prod_sec, down_sec):
    c_prod = PASTEL_COLORS['green']
    c_down = PASTEL_COLORS['red']
    
    center_text = f"<span style='font-size:18px;'>Total Run Duration</span><br><br><span style='font-size:32px; font-weight:bold; line-height:1.2'>{format_seconds_to_dhm(total_sec)}</span>"
    
    fig = go.Figure(data=[go.Pie(
        values=[prod_sec, down_sec],
        labels=['Production Time', 'Run Rate Downtime'],
        marker=dict(colors=[c_prod, c_down]),
        hole=0.7, 
        sort=False,
        direction='clockwise',
        textinfo='none',
        hoverinfo='label+percent+value'
    )])
    
    fig.update_layout(
        annotations=[dict(text=center_text, x=0.5, y=0.5, font_size=16, showarrow=False)],
        showlegend=True,
        legend=dict(orientation="h", yanchor="top", y=-0.1, xanchor="center", x=0.5, font=dict(size=14)),
        margin=dict(t=30, b=30, l=20, r=20),
        height=320,
        title=dict(text="Total Run Time Breakdown", x=0, font=dict(size=18)),
        paper_bgcolor='rgba(0,0,0,0)', 
        plot_bgcolor='rgba(0,0,0,0)'
    )
    
    fig.update_traces(textinfo='label+percent', textposition='outside', textfont=dict(size=14))
    return fig

def create_gauge(value, title, steps=None):
    """RR-style arc gauge — used for Shot Efficiency and Time Stability."""
    color = "#3498DB"
    if steps:
        if value <= 50:   color = PASTEL_COLORS['red']
        elif value <= 70: color = PASTEL_COLORS['orange']
        else:             color = PASTEL_COLORS['green']

    plot_value = max(0, min(value, 100))
    remainder  = 100 - plot_value

    fig = go.Figure(data=[go.Pie(
        values=[plot_value, remainder], hole=0.75, sort=False, direction='clockwise',
        textinfo='none', marker=dict(colors=[color, '#e6e6e6']), hoverinfo='none'
    )])
    fig.add_annotation(
        text=f"{value:.1f}%", x=0.5, y=0.5,
        font=dict(size=42, weight='bold', color=color, family="Arial"),
        showarrow=False
    )
    fig.update_layout(
        title=dict(text=title, x=0.5, xanchor='center', y=0.95, font=dict(size=16)),
        margin=dict(l=20, r=20, t=40, b=20), height=250, showlegend=False,
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)'
    )
    return fig


def create_modern_gauge(value, title):
    color = PASTEL_COLORS['green']
    if value <= 50: color = PASTEL_COLORS['red']
    elif value <= 70: color = PASTEL_COLORS['orange']
    
    plot_value = max(0, min(value, 100))
    remainder = 100 - plot_value
    visible_total = 100 
    
    values = [plot_value, remainder, visible_total]
    colors = [color, '#41424C', 'rgba(255, 255, 255, 0)']
    
    fig = go.Figure(data=[go.Pie(
        values=values,
        hole=0.65,
        sort=False,
        direction='clockwise',
        rotation=-90, 
        textinfo='none',
        marker=dict(colors=colors), 
        hoverinfo='none'
    )])

    fig.add_annotation(
        text=f"{value:.1f}%",
        x=0.5, y=0.15,
        font=dict(size=48, weight='bold', color='white', family="Arial"),
        showarrow=False
    )
    
    fig.update_layout(
        title=dict(text=title, x=0, xanchor='left', y=0.9, font=dict(size=20)),
        margin=dict(l=20, r=20, t=40, b=0),
        height=220, 
        showlegend=False,
        paper_bgcolor='rgba(0,0,0,0)', 
        plot_bgcolor='rgba(0,0,0,0)'
    )
    return fig

def create_stability_driver_bar(mtbf, mttr, stability_index):
    total = mtbf + mttr
    if total == 0: return go.Figure()
    
    mtbf_pct = (mtbf / total) * 100
    mttr_pct = 100 - mtbf_pct
    
    downtime_pct = 100 - stability_index
    label_mtbf = f"MTBF: {mtbf:.1f}m ({mtbf_pct:.1f}%)"
    label_mttr = f"MTTR: {mttr:.1f}m ({mttr_pct:.1f}%)"

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=['Cycle'], x=[mtbf], name=label_mtbf, orientation='h',
        marker_color=PASTEL_COLORS['blue'],
        hoverinfo='name' 
    ))
    fig.add_trace(go.Bar(
        y=['Cycle'], x=[mttr], name=label_mttr, orientation='h',
        marker_color=PASTEL_COLORS['red'],
        hoverinfo='name'
    ))
    
    footnote_text = f"Stability Index: {stability_index:.1f}% Stable Production Time vs. {downtime_pct:.1f}% Run Rate Downtime"

    fig.update_layout(
        barmode='stack',
        xaxis=dict(visible=False),
        yaxis=dict(visible=False),
        margin=dict(t=60, b=120, l=10, r=10), 
        height=260, 
        title=dict(text="MTTR & MTBF Analysis", x=0, font=dict(size=24)), 
        showlegend=True,
        legend=dict(
            orientation="h", 
            yanchor="top", 
            y=-0.1, 
            xanchor="center", 
            x=0.5,
            font=dict(size=18), 
            bgcolor='rgba(0,0,0,0)'
        ),
        paper_bgcolor='rgba(0,0,0,0)', 
        plot_bgcolor='rgba(0,0,0,0)',
        annotations=[
            dict(x=0, y=-0.7, text=footnote_text, showarrow=False, xref='paper', yref='paper', xanchor='left', yanchor='top', font=dict(size=16, color="#cccccc"))
        ]
    )
    return fig

def create_donut_chart(value, title, color_scheme='blue'):
    if color_scheme == 'blue': main_color = PASTEL_COLORS['blue']
    elif color_scheme == 'green': main_color = PASTEL_COLORS['green']
    elif color_scheme == 'dynamic':
        if value < 70: main_color = PASTEL_COLORS['red']
        elif value < 90: main_color = PASTEL_COLORS['orange']
        else: main_color = PASTEL_COLORS['green']
    else: main_color = color_scheme

    plot_val = min(value, 100)
    remainder = 100 - plot_val
    
    fig = go.Figure(data=[go.Pie(
        values=[plot_val, remainder], hole=0.75, sort=False, direction='clockwise',
        textinfo='none', marker=dict(colors=[main_color, '#e6e6e6']), hoverinfo='none'
    )])

    fig.add_annotation(text=f"{value:.1f}%", x=0.5, y=0.5, font=dict(size=24, weight='bold', color=main_color), showarrow=False)
    
    fig.update_layout(
        title=dict(text=title, x=0.5, xanchor='center', y=0.95, font=dict(size=14)),
        margin=dict(l=20, r=20, t=30, b=20), height=180, showlegend=False,
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)'
    )
    return fig

def plot_waterfall(metrics, benchmark_mode="Optimal"):
    total_opt = metrics['optimal_output_parts']
    actual    = metrics['actual_output_parts']
    loss_dt   = -metrics['capacity_loss_downtime_parts']
    loss_slow = -metrics['capacity_loss_slow_parts']
    gain_fast =  metrics['capacity_gain_fast_parts']

    measure   = ["absolute", "relative", "relative", "relative", "total"]
    x_label   = ["Optimal Output (parts)", "Loss: RR Downtime (parts)",
                  "Loss: Slow Cycles (parts)", "Gain: Fast Cycles (parts)", "Actual Output (parts)"]
    y_val     = [total_opt, loss_dt, loss_slow, gain_fast, actual]
    text_val  = [f"{total_opt:,.0f}", f"{loss_dt:,.0f}", f"{loss_slow:,.0f}",
                 f"+{gain_fast:,.0f}", f"{actual:,.0f}"]

    # Custom colours per bar — Plotly Waterfall colours by measure type,
    # so we use a Bar-based waterfall manually for full colour control.
    bar_colors = [
        PASTEL_COLORS['blue'],    # Optimal
        PASTEL_COLORS['red'],     # Downtime loss
        PASTEL_COLORS['orange'],  # Slow cycle loss
        PASTEL_COLORS['green'],   # Fast cycle gain
        PASTEL_COLORS['blue'],    # Actual
    ]

    # Manual bar-based waterfall for full colour control across all Plotly versions.
    # Positions: opt → (opt+loss_dt) → (opt+loss_dt+loss_slow) → actual
    pos_after_dt   = total_opt + loss_dt      # loss_dt is already negative
    pos_after_slow = pos_after_dt + loss_slow # loss_slow is already negative (== actual - gain_fast)

    bar_bases   = [0,         pos_after_dt, pos_after_slow, pos_after_slow, 0     ]
    bar_heights = [total_opt, abs(loss_dt),  abs(loss_slow), gain_fast,     actual]
    bar_texts   = [f"{total_opt:,.0f}", f"{loss_dt:,.0f}", f"{loss_slow:,.0f}",
                   f"+{gain_fast:,.0f}", f"{actual:,.0f}"]
    bar_hovers  = [
        f"Optimal Output: {total_opt:,.0f} parts<extra></extra>",
        f"Loss: RR Downtime: {abs(loss_dt):,.0f} parts<extra></extra>",
        f"Loss: Slow Cycles: {abs(loss_slow):,.0f} parts<extra></extra>",
        f"Gain: Fast Cycles: {gain_fast:,.0f} parts<extra></extra>",
        f"Actual Output: {actual:,.0f} parts<extra></extra>",
    ]

    fig = go.Figure()
    for i in range(len(x_label)):
        fig.add_trace(go.Bar(
            x=[x_label[i]],
            y=[bar_heights[i]],
            base=[bar_bases[i]],
            marker_color=bar_colors[i],
            text=[bar_texts[i]],
            textposition="outside",
            hovertemplate=bar_hovers[i],
            showlegend=False,
        ))
    fig.update_layout(barmode='overlay')

    total_target = metrics['target_output_parts']
    if "Target" in str(benchmark_mode):
        fig.add_shape(type="line", x0=-0.5, x1=4.5, y0=total_target, y1=total_target,
                      line=dict(color=PASTEL_COLORS.get('target_line', '#FFD700'),
                                width=2, dash="dash"))
        fig.add_annotation(x=0, y=total_target,
                           text=f"Target: {total_target:,.0f}",
                           showarrow=False, yshift=10)

    fig.update_layout(
        title=f"Capacity Bridge: Where am I now? (vs {benchmark_mode})",
        showlegend=False, height=450,
        yaxis_title="Output (parts)"
    )
    return fig

def plot_prediction_chart(pred_data, demand_target_total=None):
    if not pred_data: return go.Figure()

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=pred_data['historic_dates'], 
        y=pred_data['historic_cum'],
        mode='lines+markers',
        name='Actual History',
        line=dict(color=PASTEL_COLORS['blue'], width=3)
    ))

    fig.add_trace(go.Scatter(
        x=pred_data['future_dates'], 
        y=pred_data['proj_avg'],
        mode='lines',
        name=f"Forecast ({pred_data['rates']['avg']:.0f}/day)",
        line=dict(color=PASTEL_COLORS['blue'], width=2, dash='dash')
    ))

    fig.add_trace(go.Scatter(
        x=pred_data['future_dates'], 
        y=pred_data['proj_peak'],
        mode='lines',
        name=f"Best Case ({pred_data['rates']['peak']:.0f}/day)",
        line=dict(color=PASTEL_COLORS['green'], width=1, dash='dot')
    ))
    
    if pred_data['proj_req']:
        fig.add_trace(go.Scatter(
            x=pred_data['future_dates'], 
            y=pred_data['proj_req'],
            mode='lines',
            name=f"Required ({pred_data['rates']['req']:.0f}/day)",
            line=dict(color=PASTEL_COLORS['orange'], width=2, dash='longdash')
        ))
    
    if demand_target_total:
         fig.add_hline(y=demand_target_total, line_dash="solid", line_color=PASTEL_COLORS['purple'], annotation_text=f"Total Demand: {demand_target_total:,.0f}")

    fig.update_layout(
        title="Future Capacity Projection: Where will I be?",
        xaxis_title="Date",
        yaxis_title="Cumulative Output (Parts)",
        hovermode="x unified",
        height=500
    )
    return fig

def plot_performance_breakdown(df_agg, x_col, benchmark_mode):
    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=df_agg[x_col], y=df_agg['Actual Output'],
        name='Actual Output (parts)', marker_color=PASTEL_COLORS['blue']
    ))
    fig.add_trace(go.Bar(
        x=df_agg[x_col], y=df_agg['Slow Loss'],
        name='Loss: Slow Cycles (parts)', marker_color=PASTEL_COLORS['orange']
    ))
    fig.add_trace(go.Bar(
        x=df_agg[x_col], y=df_agg['Downtime Loss'],
        name='Loss: RR Downtime (parts)', marker_color=PASTEL_COLORS['red']
    ))
    fig.add_trace(go.Bar(
        x=df_agg[x_col], y=df_agg['Fast Gain'],
        name='Gain: Fast Cycles (parts)', marker_color=PASTEL_COLORS['green']
    ))
    fig.add_trace(go.Scatter(
        x=df_agg[x_col], y=df_agg['Optimal Output'],
        name='Optimal Output (parts)', mode='lines',
        line=dict(color=PASTEL_COLORS.get('optimal_line', '#FFD700'), dash='dot')
    ))
    if "Target" in str(benchmark_mode) and 'Target Output' in df_agg.columns:
        fig.add_trace(go.Scatter(
            x=df_agg[x_col], y=df_agg['Target Output'],
            name='Target Output (parts)', mode='lines',
            line=dict(color=PASTEL_COLORS.get('target_line', '#FFD700'), dash='dash')
        ))

    fig.update_layout(
        barmode='stack',
        title="Periodic Performance Breakdown",
        hovermode="x unified", height=450
    )
    return fig

def plot_shot_bar_chart(df, mode_lower=None, mode_upper=None, mode_ct=None):
    """
    CR shot bar chart with capacity-aware colour coding:
      - Blue base  = up to approved_ct for slow shots, full actual_ct for normal/fast
      - Orange top = actual_ct - approved_ct for slow cycles (loss portion)
      - Red        = actual_ct for downtime stops (time gap or hard stop)
    Approved CT shown as dashed green reference line.
    Tolerance band shown as green shaded region per run.
    """
    if df.empty:
        return None
    df = df.copy()

    _mode_y = mode_ct if isinstance(mode_ct, (int, float)) and mode_ct > 0 else (
        df['mode_ct'].mean() if 'mode_ct' in df.columns else None
    )

    df = df.sort_values('shot_time')
    _is_fast = (_mode_y is not None and _mode_y < 2.0)
    if _is_fast:
        dupes = df.groupby('shot_time').cumcount()
        df['plot_time'] = df['shot_time'] + pd.to_timedelta(dupes * 0.05, unit='s')
    else:
        df['plot_time'] = df['shot_time']

    # Bar width from median inter-shot gap
    _gaps = df['shot_time'].diff().dt.total_seconds().dropna()
    _median_gap_ms = int(_gaps.median() * 1000 * 0.8) if len(_gaps) > 0 else None
    _bw = {"width": _median_gap_ms} if _median_gap_ms else {}

    # Bar height = adj_ct_sec (true machine occupation)
    _adj = df['adj_ct_sec'] if 'adj_ct_sec' in df.columns else df['actual_ct']

    # Per-shot approved CT
    if 'approved_ct_for_run' in df.columns:
        _app = df['approved_ct_for_run'].fillna(df['approved_ct'] if 'approved_ct' in df.columns else _mode_y or 0)
    elif 'approved_ct' in df.columns:
        _app = df['approved_ct'].fillna(_mode_y or 0)
    else:
        _app = pd.Series(_mode_y or 0, index=df.index)

    # Classification — mirrors the engine's stop_flag logic:
    #   Downtime  : time-gap stop (adj != actual) or hard stop (999.9)
    #   Stop      : outside mode tolerance band (stop_flag=1, not a gap)
    #   Slow cycle: within band, actual_ct > approved_ct
    #   Fast cycle: within band, actual_ct < approved_ct
    #   Normal    : within band, actual_ct <= approved_ct (and >= approved)
    is_gap_stop   = (_adj != df['actual_ct']) | (df['actual_ct'] >= 999.9)
    is_band_stop  = (~is_gap_stop) & (df['stop_flag'] == 1)
    is_downtime   = is_gap_stop | is_band_stop
    is_slow       = (~is_downtime) & (df['actual_ct'] > _app)
    is_fast       = (~is_downtime) & (df['actual_ct'] < _app)

    # Layer heights
    df['_base']   = np.where(is_downtime, 0,
                    np.where(is_slow, 0,
                    np.where(is_fast, 0, _adj)))
    df['_orange'] = np.where(is_slow, _adj, 0)
    df['_green']  = np.where(is_fast, df['actual_ct'], 0)
    df['_red']    = np.where(is_downtime, _adj, 0)

    _slow_loss = (_adj - _app).clip(lower=0)
    _fast_gain = (_app - df['actual_ct']).clip(lower=0)

    # Unified tooltip customdata: [adj_ct, slow_loss, fast_gain]
    _cd_cols = np.column_stack([
        _adj.values,
        _slow_loss.values,
        _fast_gain.values,
    ])

    # Tooltip templates per classification
    _t_normal = (
        "<b>%{x}</b><br>"
        "<span style='color:#3498DB'>■</span> <b>Normal Shot</b><br>"
        "<b>Adj. Cycle Time:</b> %{customdata[0]:.2f}s"
        "<extra></extra>"
    )
    _t_slow = (
        "<b>%{x}</b><br>"
        "<span style='color:#F4A623'>■</span> <b>Loss: Slow Cycle</b><br>"
        "<b>Adj. Cycle Time:</b> %{customdata[0]:.2f}s<br>"
        "<b>Loss:</b> +%{customdata[1]:.2f}s above Approved CT"
        "<extra></extra>"
    )
    _t_fast = (
        "<b>%{x}</b><br>"
        "<span style='color:#77dd77'>■</span> <b>Gain: Fast Cycle</b><br>"
        "<b>Adj. Cycle Time:</b> %{customdata[0]:.2f}s<br>"
        "<b>Gain:</b> %{customdata[2]:.2f}s below Approved CT"
        "<extra></extra>"
    )
    _t_stop = (
        "<b>%{x}</b><br>"
        "<span style='color:#E74C3C'>■</span> <b>Run Rate Stop</b><br>"
        "<b>Adj. Cycle Time:</b> %{customdata[0]:.2f}s"
        "<extra></extra>"
    )

    fig = go.Figure()

    # Normal shots — full blue
    _m = ~is_downtime & ~is_slow & ~is_fast
    if _m.any():
        fig.add_trace(go.Bar(
            x=df.loc[_m, 'plot_time'], y=df.loc[_m, '_base'],
            marker_color='#3498DB', name='Normal Shot',
            customdata=_cd_cols[_m.values], hovertemplate=_t_normal, **_bw
        ))

    # Slow cycles — blue base + orange top, same tooltip on both segments
    if is_slow.any():
        fig.add_trace(go.Bar(
            x=df.loc[is_slow, 'plot_time'], y=df.loc[is_slow, '_base'],
            marker_color='#3498DB', name='Normal Shot', showlegend=False,
            customdata=_cd_cols[is_slow.values], hovertemplate=_t_slow, **_bw
        ))
        fig.add_trace(go.Bar(
            x=df.loc[is_slow, 'plot_time'], y=df.loc[is_slow, '_orange'],
            marker_color=PASTEL_COLORS['orange'], name='Loss: Slow Cycle',
            customdata=_cd_cols[is_slow.values], hovertemplate=_t_slow, **_bw
        ))

    # Fast cycles — blue bar + green stack to approved CT
    if is_fast.any():
        fig.add_trace(go.Bar(
            x=df.loc[is_fast, 'plot_time'], y=df.loc[is_fast, '_base'],
            marker_color='#3498DB', name='Normal Shot', showlegend=False,
            customdata=_cd_cols[is_fast.values], hovertemplate=_t_fast, **_bw
        ))
        fig.add_trace(go.Bar(
            x=df.loc[is_fast, 'plot_time'], y=df.loc[is_fast, '_green'],
            marker_color=PASTEL_COLORS['green'], name='Gain: Fast Cycle',
            customdata=_cd_cols[is_fast.values], hovertemplate=_t_fast, **_bw
        ))

    # Downtime — full red
    if is_downtime.any():
        fig.add_trace(go.Bar(
            x=df.loc[is_downtime, 'plot_time'], y=df.loc[is_downtime, '_red'],
            marker_color=PASTEL_COLORS['red'], name='Run Rate Stop',
            customdata=_cd_cols[is_downtime.values], hovertemplate=_t_stop, **_bw
        ))

    # Approved CT reference line
    if 'approved_ct' in df.columns:
        avg_app = float(df['approved_ct'].dropna().mean())
        if not np.isnan(avg_app):
            fig.add_hline(y=avg_app, line_dash='dash', line_color='#00FF00', line_width=1.5,
                          annotation_text=f'Approved CT: {avg_app:.1f}s',
                          annotation_position='bottom right',
                          annotation_font_color='#00FF00')

    # Tolerance band per run
    if 'mode_lower' in df.columns and 'run_id' in df.columns:
        for _, grp in df.groupby('run_id'):
            if grp.empty: continue
            fig.add_shape(type='rect', xref='x', yref='y',
                          x0=grp['shot_time'].min(), x1=grp['shot_time'].max(),
                          y0=grp['mode_lower'].iloc[0], y1=grp['mode_upper'].iloc[0],
                          fillcolor=PASTEL_COLORS['green'], opacity=0.15,
                          layer='below', line_width=0)
    elif mode_lower is not None and mode_upper is not None and not df.empty:
        fig.add_shape(type='rect', xref='x', yref='y',
                      x0=df['shot_time'].min(), x1=df['shot_time'].max(),
                      y0=mode_lower, y1=mode_upper,
                      fillcolor=PASTEL_COLORS['green'], opacity=0.15,
                      layer='below', line_width=0)

    # Run boundary lines + labels
    if 'run_id' in df.columns:
        run_starts = df.groupby('run_id')['shot_time'].min().sort_values()
        view_start = df['shot_time'].min()
        label_map = (df.drop_duplicates('run_id').set_index('run_id')['run_label'].to_dict()
                     if 'run_label' in df.columns else {})
        for i, (run_id, start_time) in enumerate(run_starts.items()):
            if i == 0 and start_time <= view_start:
                continue
            lbl = label_map.get(run_id, f'Run {i+1}')
            x_str = str(start_time)
            fig.add_shape(type='line', x0=x_str, x1=x_str, y0=0, y1=1,
                          yref='paper', line=dict(width=1.5, dash='dash', color='rgba(167,139,250,1)'))
            fig.add_annotation(x=x_str, y=0.98, yref='paper', text=lbl,
                               showarrow=False, xanchor='left',
                               font=dict(color='white', size=10, weight='bold'),
                               bgcolor='rgba(60,0,90,0.80)',
                               bordercolor='rgba(167,139,250,1)', borderwidth=1, borderpad=3)

    # Y-axis: 99th percentile so outliers don't destroy scale
    _cts = df['actual_ct'].dropna()
    _p99 = float(np.percentile(_cts, 99)) if len(_cts) > 0 else 200
    y_cap = max(_p99 * 1.2, (_mode_y or 50) * 1.5)

    fig.update_layout(
        barmode='stack',
        title='Cycle Time Analysis', xaxis_title='Date / Time',
        yaxis_title='Cycle Time (sec)', yaxis=dict(range=[0, y_cap]),
        bargap=0.05, xaxis=dict(showgrid=True),
        hoverlabel=dict(bgcolor='#1e1e2e', font_size=13, font_family='monospace'),
        legend=dict(title='Legend', orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1)
    )
    return fig


def plot_shot_analysis(df_shots, zoom_y=None):
    if df_shots.empty: return go.Figure()
    fig = go.Figure()
    color_map = {'Slow Cycle': PASTEL_COLORS['red'], 'Fast Cycle': PASTEL_COLORS['orange'], 'On Target': PASTEL_COLORS['blue'], 'Downtime (Stop)': PASTEL_COLORS['grey'], 'Run Break (Excluded)': '#d3d3d3'}
    
    for shot_type, color in color_map.items():
        subset = df_shots[df_shots['shot_type'] == shot_type]
        if not subset.empty:
            fig.add_trace(go.Bar(x=subset['shot_time'], y=subset['actual_ct'], name=shot_type, marker_color=color, hovertemplate='Time: %{x}<br>CT: %{y:.2f}s<extra></extra>'))
            
    if 'run_id' in df_shots.columns:
        run_starts = df_shots.groupby('run_id')['shot_time'].min().sort_values()
        
        for i, start_time in enumerate(run_starts):
            if i > 0: 
                 fig.add_vline(x=start_time.timestamp() * 1000, line_width=2, line_dash="dash", line_color="purple")

    for r_id, run_df in df_shots.groupby('run_id'):
        lower = run_df['mode_lower'].iloc[0]
        upper = run_df['mode_upper'].iloc[0]
        start = run_df['shot_time'].min()
        end = run_df['shot_time'].max()
        
        fig.add_shape(type="rect", x0=start, x1=end, y0=lower, y1=upper, fillcolor="grey", opacity=0.2, line_width=0)
        if r_id == 0: fig.add_annotation(x=start, y=upper, text="Mode Tolerance Band", showarrow=False, yshift=10, font=dict(color="grey", size=10))

    avg_ref = df_shots['approved_ct'].mean()
    fig.add_hline(y=avg_ref, line_dash="dash", line_color="green", annotation_text=f"Avg Approved CT: {avg_ref:.2f}s")
    
    if zoom_y is None and not df_shots.empty:
        cts = df_shots['actual_ct']
        if len(cts) > 0:
            fig.update_yaxes(range=[0, min(cts.max() * 1.1, 1000)])

    fig.update_layout(
        title="Cycle Time Analysis",
        yaxis_title="Cycle Time (s)",
        xaxis_title="Time",
        barmode='overlay',
        hovermode="x unified",
        height=500,
        showlegend=True
    )
    return fig


# --- TMD LOG ---
# ==============================================================================

def load_tmd_log(file) -> pd.DataFrame:
    """
    Loads a TMD log file. Expected columns:
        Date/Time   — session event timestamp
        Machine ID  — press identifier
        Tooling ID  — tool identifier (blank on UNMATCHED)
        Status      — AUTO_MATCHED | UNMATCHED
    Returns a cleaned DataFrame sorted by Machine ID, Date/Time.
    """
    try:
        df = pd.read_excel(file) if hasattr(file, 'name') and file.name.endswith(('.xls','.xlsx')) else pd.read_csv(file)
        df.columns = [c.strip() for c in df.columns]

        col_map = {c.strip().upper(): c for c in df.columns}
        def _gc(*targets):
            for t in targets:
                found = col_map.get(t.upper())
                if found: return found
            return None

        dt_col   = _gc("DATE/TIME","DATETIME","DATE TIME","TIMESTAMP","DATE")
        mach_col = _gc("MACHINE ID","MACHINE_ID","MACHINE")
        tool_col = _gc("TOOLING ID","TOOLING_ID","TOOL ID","TOOL_ID","EQUIPMENT_CODE")
        stat_col = _gc("STATUS")

        rename = {}
        if dt_col   and dt_col   != "Date/Time":   rename[dt_col]   = "Date/Time"
        if mach_col and mach_col != "Machine ID":   rename[mach_col] = "Machine ID"
        if tool_col and tool_col != "Tooling ID":   rename[tool_col] = "Tooling ID"
        if stat_col and stat_col != "Status":       rename[stat_col] = "Status"
        df.rename(columns=rename, inplace=True)

        df["Date/Time"] = pd.to_datetime(df["Date/Time"], errors="coerce")
        df.dropna(subset=["Date/Time","Machine ID","Status"], inplace=True)
        df["Tooling ID"] = df["Tooling ID"].fillna("").astype(str).str.strip()
        df["Status"]     = df["Status"].str.strip().str.upper()
        df = df[df["Status"].isin(["AUTO_MATCHED","UNMATCHED"])].copy()
        return df.sort_values(["Machine ID","Date/Time"]).reset_index(drop=True)
    except Exception as e:
        return pd.DataFrame()


def assign_machine_from_tmd(df_shots: pd.DataFrame, df_tmd: pd.DataFrame,
                            shift_config: list = None) -> pd.DataFrame:
    """
    Joins shots to TMD sessions by timestamp window.
    Adds: machine_id, session_id, session_period (from shift_config or default 3-shift).
    shift_config: list of (name, start_hr, end_hr) tuples.
    """
    if df_tmd.empty or df_shots.empty:
        return df_shots

    # Default: 3 shifts from 06:00
    if not shift_config:
        shift_config = [("Shift 1", 6, 14), ("Shift 2", 14, 22), ("Shift 3", 22, 6)]

    def _period(dt):
        h = dt.hour
        for name, start, end in shift_config:
            if start < end:
                if start <= h < end:
                    return name
            else:  # wraps midnight
                if h >= start or h < end:
                    return name
        return shift_config[0][0]

    df_shots = df_shots.copy()
    df_shots["machine_id"]     = None
    df_shots["session_id"]     = None
    df_shots["session_period"] = None

    session_counter = 0
    for machine_id, grp in df_tmd.groupby("Machine ID"):
        grp = grp.sort_values("Date/Time").reset_index(drop=True)
        matched   = grp[grp["Status"] == "AUTO_MATCHED"].reset_index(drop=True)
        unmatched = grp[grp["Status"] == "UNMATCHED"].reset_index(drop=True)

        for _, m_row in matched.iterrows():
            session_start  = m_row["Date/Time"]
            tool_id        = m_row["Tooling ID"]
            following      = unmatched[unmatched["Date/Time"] > session_start]
            session_end    = following.iloc[0]["Date/Time"] if not following.empty else df_shots["shot_time"].max()
            session_counter += 1
            sid    = f"S{session_counter:04d}"
            period = _period(session_start)

            mask = (
                (df_shots["shot_time"] >= session_start) &
                (df_shots["shot_time"] <  session_end) &
                (df_shots["tool_id"].astype(str) == str(tool_id))
            )
            df_shots.loc[mask, "machine_id"]     = machine_id
            df_shots.loc[mask, "session_id"]     = sid
            df_shots.loc[mask, "session_period"] = period

    return df_shots


# ==============================================================================
# --- MACHINE FIT ANALYSIS ---
# ==============================================================================

def compute_machine_fit_metrics(df_processed: pd.DataFrame, config: dict) -> pd.DataFrame:
    """
    Computes per-(tool_id, machine_id) performance metrics from the already-processed
    dataframe (stop_flag, run_id, approved_ct_for_run etc. already set by
    CapacityRiskCalculator). Returns one row per tool-machine pair with a composite
    fit score (0–100, higher = better match).
    """
    if 'machine_id' not in df_processed.columns:
        return pd.DataFrame()

    records = []
    for (tool_id, machine_id), grp in df_processed.groupby(['tool_id', 'machine_id']):
        if grp.empty:
            continue

        n_runs      = grp['run_id'].nunique()
        total_shots = len(grp)
        normal_shots = int((grp['stop_flag'] == 0).sum())
        stop_events  = int(grp['stop_event'].sum()) if 'stop_event' in grp.columns else 0

        prod_df   = grp[grp['stop_flag'] == 0]
        prod_time = float(prod_df['actual_ct'].sum())

        total_runtime = 0.0
        opt_output    = 0.0
        for _, run_df in grp.groupby('run_id'):
            if run_df.empty:
                continue
            dur = (run_df['shot_time'].max() - run_df['shot_time'].min()).total_seconds() + float(run_df.iloc[-1]['actual_ct'])
            total_runtime += dur
            r_ct  = float(run_df['approved_ct_for_run'].iloc[0]) if 'approved_ct_for_run' in run_df.columns else float(run_df['approved_ct'].iloc[0])
            r_cav = float(run_df['working_cavities'].max()) if 'working_cavities' in run_df.columns else 1.0
            if r_ct > 0:
                opt_output += (dur / r_ct) * r_cav

        downtime   = max(0.0, total_runtime - prod_time)
        act_output = float(prod_df['working_cavities'].sum()) if 'working_cavities' in prod_df.columns else float(normal_shots)

        cap_gain_fast = cap_loss_slow = 0.0
        if not prod_df.empty and 'approved_ct_for_run' in prod_df.columns:
            parts_delta = (
                (prod_df['approved_ct_for_run'] - prod_df['actual_ct'])
                / prod_df['approved_ct_for_run'].replace(0, np.nan)
            ) * (prod_df['working_cavities'] if 'working_cavities' in prod_df.columns else 1)
            cap_gain_fast = float(parts_delta[parts_delta > 0].sum())
            cap_loss_slow = float(abs(parts_delta[parts_delta < 0].sum()))

        eff_rate  = (normal_shots / total_shots  * 100) if total_shots  > 0 else 0.0
        stab_idx  = (prod_time   / total_runtime * 100) if total_runtime > 0 else 0.0
        cap_eff   = (act_output  / opt_output    * 100) if opt_output   > 0 else 0.0
        mttr_min  = (downtime / 60 / stop_events)       if stop_events  > 0 else 0.0
        mtbf_min  = (prod_time / 60 / stop_events)      if stop_events  > 0 else (prod_time / 60)

        avg_ct = prod_df['actual_ct'].mean() if not prod_df.empty else 0.0
        ct_cv  = (prod_df['actual_ct'].std() / avg_ct * 100) if (not prod_df.empty and avg_ct > 0) else 0.0

        # Improvement rate: compare cap efficiency of first-half runs vs second-half runs
        run_ids_sorted = grp.groupby('run_id')['shot_time'].min().sort_values().index.tolist()
        improvement_rate = np.nan
        if len(run_ids_sorted) >= 2:
            mid = len(run_ids_sorted) // 2
            first_half = run_ids_sorted[:mid]
            second_half = run_ids_sorted[mid:]
            def _run_cap_eff(run_ids):
                sub = grp[grp['run_id'].isin(run_ids)]
                sub_prod = sub[sub['stop_flag'] == 0]
                sub_act = float(sub_prod['working_cavities'].sum()) if 'working_cavities' in sub_prod.columns else float(len(sub_prod))
                sub_opt = 0.0
                for _, r in sub.groupby('run_id'):
                    dur = (r['shot_time'].max() - r['shot_time'].min()).total_seconds() + float(r.iloc[-1]['actual_ct'])
                    r_ct = float(r['approved_ct_for_run'].iloc[0]) if 'approved_ct_for_run' in r.columns else float(r['approved_ct'].iloc[0])
                    r_cav = float(r['working_cavities'].max()) if 'working_cavities' in r.columns else 1.0
                    if r_ct > 0: sub_opt += (dur / r_ct) * r_cav
                return (sub_act / sub_opt * 100) if sub_opt > 0 else 0.0
            eff_first  = _run_cap_eff(first_half)
            eff_second = _run_cap_eff(second_half)
            improvement_rate = round(eff_second - eff_first, 1)

        records.append({
            'tool_id':             tool_id,
            'machine_id':          machine_id,
            'supplier_id':         str(grp['supplier_id'].iloc[0]) if 'supplier_id' in grp.columns else 'Unknown',
            'runs':                n_runs,
            'total_shots':         total_shots,
            'production_hrs':      round(total_runtime / 3600, 1),
            'total_parts':         round(act_output, 0),
            'avg_ct_sec':          round(avg_ct, 2),
            'ct_fluctuation_pct':  round(ct_cv, 2),
            'efficiency_pct':      round(eff_rate, 1),
            'stability_pct':       round(stab_idx, 1),
            'cap_efficiency_pct':  round(cap_eff, 1),
            'improvement_rate':    improvement_rate,
            'slow_loss_parts':     round(cap_loss_slow, 0),
            'fast_gain_parts':     round(cap_gain_fast, 0),
            'mtbf_min':            round(mtbf_min, 1),
            'mttr_min':            round(mttr_min, 1),
            'stop_count':          stop_events,
            'downtime_hrs':        round(downtime / 3600, 2),
            'opt_output':          round(opt_output, 0),
            'act_output':          round(act_output, 0),
        })

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)

    # Composite fit score (0–100): normalised per-tool so machines are ranked
    # relative to each other for that specific tool.
    score_rows = []
    for _, tgrp in df.groupby('tool_id'):
        def _hi(s): return (s - s.min()) / (s.max() - s.min() + 1e-9)
        def _lo(s): return 1.0 - _hi(s)
        score = (
            _hi(tgrp['cap_efficiency_pct']) * 30 +
            _hi(tgrp['stability_pct'])       * 25 +
            _hi(tgrp['efficiency_pct'])      * 20 +
            _hi(tgrp['mtbf_min'])            * 15 +
            _lo(tgrp['ct_fluctuation_pct'])  * 10
        )
        score_rows.append(score)

    df['fit_score'] = pd.concat(score_rows).reindex(df.index).round(1)

    # Rename ct_cv_pct legacy ref if any — already stored as ct_fluctuation_pct
    return df


def compute_supplier_scorecard(fit_df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregates fit_df by supplier_id for the holistic supply chain overview.
    Shows which suppliers have the best machine-tool fit performance.
    Returns one row per supplier, ranked by avg fit score.
    """
    if 'supplier_id' not in fit_df.columns:
        return pd.DataFrame()

    agg = fit_df.groupby('supplier_id').agg(
        total_tools       =('tool_id',            'nunique'),
        total_machines    =('machine_id',          'nunique'),
        total_runs        =('runs',               'sum'),
        total_parts       =('total_parts',         'sum'),
        production_hrs    =('production_hrs',      'sum'),
        avg_fit_score     =('fit_score',           'mean'),
        avg_cap_eff       =('cap_efficiency_pct',  'mean'),
        avg_stability     =('stability_pct',       'mean'),
        avg_efficiency    =('efficiency_pct',      'mean'),
        avg_mtbf          =('mtbf_min',            'mean'),
        avg_mttr          =('mttr_min',            'mean'),
        total_slow_loss   =('slow_loss_parts',     'sum'),
        total_fast_gain   =('fast_gain_parts',     'sum'),
    ).round(1).reset_index()

    # Best and worst machine per supplier by avg fit score across their tools
    best_mach = fit_df.groupby(['supplier_id','machine_id'])['fit_score'].mean() \
                      .reset_index().sort_values('fit_score', ascending=False)
    agg['best_machine']  = agg['supplier_id'].map(
        best_mach.groupby('supplier_id').first()['machine_id'])
    agg['worst_machine'] = agg['supplier_id'].map(
        best_mach.groupby('supplier_id').last()['machine_id'])

    agg['rank'] = agg['avg_fit_score'].rank(ascending=False, method='min').astype(int)
    return agg.sort_values('rank').reset_index(drop=True)


def plot_machine_fit_heatmap(fit_df: pd.DataFrame) -> go.Figure:
    """
    Heatmap of machine-tool performance. Machines on Y, metrics on X.
    Each cell normalised within the column (0–1); green = better, red = worse.
    Direction (higher/lower is better) is handled per metric.
    NaN values (e.g. improvement_rate with only 1 run) shown as neutral grey.
    """
    # (column, display label, True=higher better, False=lower better, None=neutral)
    METRIC_CFG = [
        ('fit_score',            'Fit Score',       True),
        ('cap_efficiency_pct',   'Cap Eff %',       True),
        ('stability_pct',        'Stability %',     True),
        ('efficiency_pct',       'Efficiency %',    True),
        ('improvement_rate',     'Improvement',     True),
        ('mtbf_min',             'MTBF (min)',      True),
        ('mttr_min',             'MTTR (min)',      False),
        ('slow_loss_parts',      'Slow Loss',       False),
        ('fast_gain_parts',      'Fast Gain',       True),
        ('ct_fluctuation_pct',   'CT Fluctuation%', False),
        ('avg_ct_sec',           'Avg CT (s)',      None),
        ('production_hrs',       'Prod Hrs',        None),
        ('total_parts',          'Total Parts',     None),
        ('runs',                 'Runs',            None),
    ]

    machines = fit_df['machine_id'].tolist()
    z_norm, z_text, col_labels = [], [], []

    for col, label, direction in METRIC_CFG:
        if col not in fit_df.columns:
            continue
        vals = fit_df[col].values.astype(float)
        valid = vals[~np.isnan(vals)]
        vmin = valid.min() if len(valid) else 0.0
        vmax = valid.max() if len(valid) else 0.0
        rng  = vmax - vmin

        norm = np.full_like(vals, 0.5, dtype=float)  # NaN or neutral → 0.5
        non_nan = ~np.isnan(vals)
        if rng >= 1e-9 and direction is not None:
            if direction:
                norm[non_nan] = (vals[non_nan] - vmin) / rng
            else:
                norm[non_nan] = 1.0 - (vals[non_nan] - vmin) / rng

        z_norm.append(norm)
        z_text.append([f"{v:,.1f}" if not np.isnan(v) else "—" for v in vals])
        col_labels.append(label)

    z  = np.array(z_norm).T   # shape: (n_machines, n_metrics)
    zt = np.array(z_text).T

    fig = go.Figure(go.Heatmap(
        z=z,
        x=col_labels,
        y=machines,
        colorscale=[[0, PASTEL_COLORS['red']], [0.5, '#FFFFAA'], [1, PASTEL_COLORS['green']]],
        zmin=0, zmax=1,
        text=zt,
        texttemplate="%{text}",
        customdata=zt,
        hovertemplate="%{y}  |  %{x}: %{customdata}<extra></extra>",
        showscale=False,
    ))
    fig.update_layout(
        title="Performance Heatmap — Green = Better for that metric",
        xaxis_title="Metric",
        yaxis_title="Machine",
        height=max(320, 52 * len(machines)),
        margin=dict(l=120, r=20, t=60, b=60),
        xaxis=dict(side='top'),
    )
    return fig


def compute_machine_tool_rankings(fit_df: pd.DataFrame) -> pd.DataFrame:
    """
    Machine-centric view: for each machine, rank all tools that have run on it.
    Adds columns:
      rank_on_machine    — 1 = best tool on this machine
      vs_best_pct        — cap efficiency delta vs the best tool on this machine (negative = worse)
      vs_worst_pct       — cap efficiency delta vs the worst tool on this machine (positive = gain potential)
      parts_gain_potential — extra parts per production hour if this machine ran the best tool instead
    """
    if fit_df.empty:
        return pd.DataFrame()

    out = []
    for machine_id, grp in fit_df.groupby('machine_id'):
        grp = grp.copy().sort_values('cap_efficiency_pct', ascending=False).reset_index(drop=True)
        best_eff  = grp['cap_efficiency_pct'].max()
        worst_eff = grp['cap_efficiency_pct'].min()

        # Parts per hour for the best tool on this machine
        best_row = grp.iloc[0]
        best_parts_per_hr = (best_row['total_parts'] / best_row['production_hrs']
                             if best_row['production_hrs'] > 0 else 0)

        grp['rank_on_machine']     = range(1, len(grp) + 1)
        grp['vs_best_pct']         = (grp['cap_efficiency_pct'] - best_eff).round(1)
        grp['vs_worst_pct']        = (grp['cap_efficiency_pct'] - worst_eff).round(1)
        # Parts gain potential: how many more parts/hr this machine would produce
        # if the best tool ran instead of this tool
        grp['parts_gain_potential'] = grp.apply(
            lambda r: round((best_parts_per_hr - (r['total_parts'] / r['production_hrs']
                             if r['production_hrs'] > 0 else 0)), 1),
            axis=1
        )
        out.append(grp)

    return pd.concat(out, ignore_index=True) if out else pd.DataFrame()


def compute_recommendations(fit_df: pd.DataFrame) -> pd.DataFrame:
    """
    Generates best-match recommendations and cross-supplier swap opportunities.

    Returns a DataFrame with one row per machine showing:
      - Current best tool on that machine
      - If a tool from a different supplier would outperform — a swap recommendation
      - Quantified gain: extra parts/hr and cap efficiency gain
    """
    if fit_df.empty or 'supplier_id' not in fit_df.columns:
        return pd.DataFrame()

    recs = []
    for machine_id, grp in fit_df.groupby('machine_id'):
        grp = grp.sort_values('cap_efficiency_pct', ascending=False).reset_index(drop=True)
        if grp.empty:
            continue

        best = grp.iloc[0]
        worst = grp.iloc[-1] if len(grp) > 1 else best

        # Parts per hour for each tool
        def _pph(row):
            return row['total_parts'] / row['production_hrs'] if row['production_hrs'] > 0 else 0

        best_pph  = _pph(best)
        worst_pph = _pph(worst)

        # Cross-supplier opportunity: best tool from a DIFFERENT supplier than the worst tool
        cross = grp[grp['supplier_id'] != worst['supplier_id']]
        swap_tool = cross.iloc[0] if not cross.empty else None

        rec = {
            'machine_id':           machine_id,
            'best_tool':            best['tool_id'],
            'best_supplier':        best['supplier_id'],
            'best_cap_eff':         round(best['cap_efficiency_pct'], 1),
            'best_parts_per_hr':    round(best_pph, 1),
            'worst_tool':           worst['tool_id'],
            'worst_supplier':       worst['supplier_id'],
            'worst_cap_eff':        round(worst['cap_efficiency_pct'], 1),
            'worst_parts_per_hr':   round(worst_pph, 1),
            'tools_compared':       len(grp),
            'cap_eff_spread':       round(best['cap_efficiency_pct'] - worst['cap_efficiency_pct'], 1),
            'cap_eff_gain':         round(best['cap_efficiency_pct'] - worst['cap_efficiency_pct'], 1),
        }

        if swap_tool is not None and swap_tool['tool_id'] != best['tool_id']:
            swap_pph = _pph(swap_tool)
            # Find current tool from swap_tool's supplier on this machine
            current_same_sup = grp[grp['supplier_id'] == swap_tool['supplier_id']]
            current_pph = _pph(current_same_sup.iloc[-1]) if not current_same_sup.empty else worst_pph
            rec.update({
                'swap_recommended':      True,
                'swap_tool':             swap_tool['tool_id'],
                'swap_from_supplier':    worst['supplier_id'],
                'swap_to_supplier':      swap_tool['supplier_id'],
                'swap_cap_eff_gain':     round(swap_tool['cap_efficiency_pct'] - worst['cap_efficiency_pct'], 1),
                'swap_parts_per_hr_gain':round(swap_pph - current_pph, 1),
            })
        else:
            rec.update({
                'swap_recommended':      False,
                'swap_tool':             '—',
                'swap_from_supplier':    '—',
                'swap_to_supplier':      '—',
                'swap_cap_eff_gain':     0.0,
                'swap_parts_per_hr_gain':0.0,
            })

        recs.append(rec)

    return pd.DataFrame(recs).sort_values('cap_eff_spread', ascending=False).reset_index(drop=True)


def compute_part_recommendations(fit_df: pd.DataFrame, copy_map: dict = None) -> pd.DataFrame:
    """
    Part/tool-centric view: for each part (copy group) or individual tool,
    finds the optimal machine based on historical Cap Efficiency.

    Returns one row per part, sorted by cap_eff_gain descending (largest
    opportunity first). Each row carries aggregated machine-level stats for
    the best and worst machines so the UI can render without re-filtering.
    """
    if fit_df.empty:
        return pd.DataFrame()

    # Build part_id → [tool_ids] mapping
    if copy_map:
        part_to_tools: dict = {}
        for tool_id, part in copy_map.items():
            if tool_id in fit_df['tool_id'].values:
                part_to_tools.setdefault(part, []).append(tool_id)
        # Tools not assigned to any copy group → individual entries
        for t in fit_df['tool_id'].unique():
            if t not in copy_map:
                part_to_tools[t] = [t]
    else:
        part_to_tools = {t: [t] for t in fit_df['tool_id'].unique()}

    recs = []
    for part_id, tool_ids in part_to_tools.items():
        part_fit = fit_df[fit_df['tool_id'].isin(tool_ids)]
        if part_fit.empty:
            continue

        # Aggregate by machine: average across all copy tools of this part
        by_machine = (
            part_fit.groupby('machine_id')
            .agg(
                avg_cap_eff    =('cap_efficiency_pct', 'mean'),
                avg_stability  =('stability_pct',      'mean'),
                avg_mtbf       =('mtbf_min',           'mean'),
                avg_mttr       =('mttr_min',           'mean'),
                total_parts    =('total_parts',        'sum'),
                production_hrs =('production_hrs',     'sum'),
                runs           =('runs',               'sum'),
            )
            .round(1)
            .reset_index()
            .sort_values('avg_cap_eff', ascending=False)
            .reset_index(drop=True)
        )

        if by_machine.empty:
            continue

        best  = by_machine.iloc[0]
        worst = by_machine.iloc[-1]

        # Supplier of the tool with highest cap eff on the best machine
        best_tool_rows = part_fit[part_fit['machine_id'] == best['machine_id']] \
                             .sort_values('cap_efficiency_pct', ascending=False)
        best_supplier = best_tool_rows.iloc[0]['supplier_id'] if not best_tool_rows.empty else '—'
        best_tool_id  = best_tool_rows.iloc[0]['tool_id']     if not best_tool_rows.empty else '—'

        recs.append({
            'part_id':         part_id,
            'tools':           ', '.join(sorted(tool_ids)),
            'tool_count':      len(tool_ids),
            'machines_tested': len(by_machine),
            'best_machine':    best['machine_id'],
            'best_tool_id':    best_tool_id,
            'best_supplier':   best_supplier,
            'best_cap_eff':    round(best['avg_cap_eff'],   1),
            'best_stability':  round(best['avg_stability'], 1),
            'best_mtbf':       round(best['avg_mtbf'],      1),
            'best_mttr':       round(best['avg_mttr'],      1),
            'best_prod_hrs':   round(best['production_hrs'],1),
            'best_parts':      round(best['total_parts'],   0),
            'worst_machine':   worst['machine_id'],
            'worst_cap_eff':   round(worst['avg_cap_eff'],  1),
            'avg_cap_eff':     round(by_machine['avg_cap_eff'].mean(), 1),
            'cap_eff_gain':    round(best['avg_cap_eff'] - worst['avg_cap_eff'], 1),
        })

    if not recs:
        return pd.DataFrame()
    return (pd.DataFrame(recs)
              .sort_values('cap_eff_gain', ascending=False)
              .reset_index(drop=True))


# ==============================================================================
# --- MATCH EFFICIENCY RATE ---
# ==============================================================================

def compute_match_efficiency_rate(fit_df: pd.DataFrame, threshold: float = 85.0) -> pd.DataFrame:
    """
    Per supplier: % of machine-tool sessions where cap_efficiency_pct >= threshold.
    Also returns avg cap eff, total sessions, parts, prod hrs.
    """
    if fit_df.empty or 'supplier_id' not in fit_df.columns:
        return pd.DataFrame()

    rows = []
    for sup, grp in fit_df.groupby('supplier_id'):
        total     = len(grp)
        efficient = (grp['cap_efficiency_pct'] >= threshold).sum()
        rows.append({
            'supplier_id':          sup,
            'total_sessions':       total,
            'efficient_sessions':   int(efficient),
            'match_efficiency_pct': round(efficient / total * 100, 1) if total > 0 else 0.0,
            'avg_cap_eff':          round(grp['cap_efficiency_pct'].mean(), 1),
            'total_parts':          round(grp['total_parts'].sum(), 0),
            'production_hrs':       round(grp['production_hrs'].sum(), 1),
            'tools':                grp['tool_id'].nunique(),
            'machines_used':        grp['machine_id'].nunique(),
        })
    return pd.DataFrame(rows).sort_values('match_efficiency_pct', ascending=False).reset_index(drop=True)


# ==============================================================================
# --- PRESS COMPARE ---
# ==============================================================================

def compute_press_compare(fit_df: pd.DataFrame, tool_ids: list,
                           df_processed: pd.DataFrame = None,
                           recent_days: int = 30) -> dict:
    """
    Tool-driven press comparison. Given one or more tool_ids (a part or copy group):
    - All-time metrics per machine
    - Recent (last N days) vs historical split
    - % delta vs group average for each KPI
    Returns dict with keys: 'alltime', 'recent', 'historical', 'delta'
    """
    KPIS = ['cap_efficiency_pct', 'stability_pct', 'efficiency_pct',
            'mtbf_min', 'mttr_min', 'avg_ct_sec', 'total_parts',
            'production_hrs', 'slow_loss_parts', 'fast_gain_parts',
            'stop_count', 'fit_score']

    KPI_LABELS = {
        'cap_efficiency_pct': 'Cap Efficiency %',
        'stability_pct':      'Stability %',
        'efficiency_pct':     'Efficiency %',
        'mtbf_min':           'MTBF (min)',
        'mttr_min':           'MTTR (min)',
        'avg_ct_sec':         'Avg CT (s)',
        'total_parts':        'Parts Produced',
        'production_hrs':     'Prod Hours',
        'slow_loss_parts':    'Slow Cycle Loss',
        'fast_gain_parts':    'Fast Cycle Gain',
        'stop_count':         'Stop Events',
        'fit_score':          'Fit Score',
    }
    # Higher = better for these; lower = better for rest
    HIGHER_BETTER = {'cap_efficiency_pct','stability_pct','efficiency_pct',
                     'mtbf_min','total_parts','production_hrs','fast_gain_parts','fit_score'}

    base = fit_df[fit_df['tool_id'].isin(tool_ids)].copy()
    if base.empty:
        return {}

    # All-time: pivot machine × KPI
    alltime = base.groupby('machine_id')[KPIS].mean().round(2)

    # % delta vs group mean
    group_mean = alltime.mean()
    delta = alltime.copy()
    for col in KPIS:
        if group_mean[col] != 0:
            delta[col] = ((alltime[col] - group_mean[col]) / group_mean[col] * 100).round(1)
        else:
            delta[col] = 0.0

    # Recent vs historical split using session timestamps from df_processed
    recent_df = hist_df = pd.DataFrame()
    if df_processed is not None and not df_processed.empty and 'session_id' in df_processed.columns:
        sub = df_processed[df_processed['tool_id'].isin(tool_ids)].copy()
        if not sub.empty:
            cutoff = sub['shot_time'].max() - pd.Timedelta(days=recent_days)
            sub['is_recent'] = sub['shot_time'] >= cutoff

            def _agg(mask):
                g = sub[mask].groupby('machine_id')
                out = []
                for m, mg in g:
                    prod = mg[mg['stop_flag'] == 0]
                    prod_time = float(prod['actual_ct'].sum())
                    total_rt  = (mg['shot_time'].max() - mg['shot_time'].min()).total_seconds() + float(mg.iloc[-1]['actual_ct'])
                    stops = int(mg['stop_event'].sum()) if 'stop_event' in mg.columns else 0
                    act_out = float(prod['working_cavities'].sum()) if 'working_cavities' in prod.columns else float(len(prod))
                    opt_out_s = 0.0
                    for _, rg in mg.groupby('run_id'):
                        dur = (rg['shot_time'].max()-rg['shot_time'].min()).total_seconds()+float(rg.iloc[-1]['actual_ct'])
                        rct = float(rg['approved_ct_for_run'].iloc[0]) if 'approved_ct_for_run' in rg.columns else float(rg['approved_ct'].iloc[0])
                        rc  = float(rg['working_cavities'].max()) if 'working_cavities' in rg.columns else 1.0
                        if rct > 0: opt_out_s += (dur/rct)*rc
                    out.append({
                        'machine_id':        m,
                        'cap_efficiency_pct':round(act_out/opt_out_s*100,1) if opt_out_s>0 else 0,
                        'stability_pct':     round(prod_time/total_rt*100,1) if total_rt>0 else 0,
                        'stop_count':        stops,
                        'total_parts':       round(act_out,0),
                        'production_hrs':    round(total_rt/3600,1),
                        'mtbf_min':          round(prod_time/60/stops,1) if stops>0 else round(prod_time/60,1),
                        'mttr_min':          round((total_rt-prod_time)/60/stops,1) if stops>0 else 0,
                    })
                return pd.DataFrame(out).set_index('machine_id') if out else pd.DataFrame()

            recent_df = _agg(sub['is_recent'])
            hist_df   = _agg(~sub['is_recent'])

    return {
        'alltime':     alltime,
        'delta':       delta,
        'recent':      recent_df,
        'historical':  hist_df,
        'kpi_labels':  KPI_LABELS,
        'higher_better': HIGHER_BETTER,
    }


# ==============================================================================
# --- WEEKLY REPORT GENERATOR ---
# ==============================================================================

def generate_weekly_report(fit_df: pd.DataFrame,
                            scorecard: pd.DataFrame,
                            recs_df: pd.DataFrame,
                            mer_df: pd.DataFrame,
                            rankings_df: pd.DataFrame,
                            report_date: str = None) -> bytes:
    """
    Generates a weekly Excel report as bytes (for st.download_button).
    Sheets: Summary, Match Efficiency, Supplier Scorecard,
            Optimal Pairings, Machine Rankings, All Sessions.
    """
    import io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
    RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
    AMB_FILL    = PatternFill("solid", fgColor="FFEB9C")

    def _write_sheet(ws, df, title=None):
        if title:
            ws.append([title])
            ws.cell(1,1).font = Font(bold=True, size=13)
            ws.append([])
        if df.empty:
            ws.append(["No data available"])
            return
        ws.append(list(df.columns))
        hrow = ws.max_row
        for cell in ws[hrow]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
        for row in df.itertuples(index=False):
            ws.append(list(row))
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    def _color_col(ws, col_idx, good_thresh, bad_thresh, higher_better=True, data_start_row=3):
        for row in ws.iter_rows(min_row=data_start_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    v = float(cell.value)
                    if higher_better:
                        cell.fill = GREEN_FILL if v >= good_thresh else (RED_FILL if v < bad_thresh else AMB_FILL)
                    else:
                        cell.fill = GREEN_FILL if v <= good_thresh else (RED_FILL if v > bad_thresh else AMB_FILL)
                except (TypeError, ValueError):
                    pass

    from openpyxl import Workbook
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    date_str = report_date or pd.Timestamp.now().strftime("%Y-%m-%d")
    ws_sum.append([f"Machine Fit Weekly Report — {date_str}"])
    ws_sum.cell(1,1).font = Font(bold=True, size=14)
    ws_sum.append([])
    ws_sum.append(["Metric", "Value"])
    for c in ws_sum[3]: c.fill=HEADER_FILL; c.font=HEADER_FONT
    if not fit_df.empty:
        ws_sum.append(["Suppliers Tracked",      fit_df['supplier_id'].nunique() if 'supplier_id' in fit_df.columns else '—'])
        ws_sum.append(["Tools Tracked",           fit_df['tool_id'].nunique()])
        ws_sum.append(["Total Sessions",          len(fit_df)])
        ws_sum.append(["Total Parts Produced",    f"{fit_df['total_parts'].sum():,.0f}"])
        ws_sum.append(["Total Production Hours",  f"{fit_df['production_hrs'].sum():,.1f}"])
        ws_sum.append(["Avg Cap Efficiency",      f"{fit_df['cap_efficiency_pct'].mean():.1f}%"])
        ws_sum.append(["Avg Fit Score",           f"{fit_df['fit_score'].mean():.1f} / 100"])
    if not mer_df.empty:
        top_sup = mer_df.iloc[0]
        ws_sum.append(["Top Supplier (Match Eff)", f"{top_sup['supplier_id']} — {top_sup['match_efficiency_pct']:.1f}%"])
    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 25

    # ── Sheet 2: Match Efficiency ─────────────────────────────────────────────
    ws_mer = wb.create_sheet("Match Efficiency")
    _write_sheet(ws_mer, mer_df.rename(columns={
        'supplier_id':'Supplier','total_sessions':'Total Sessions',
        'efficient_sessions':'Efficient Sessions','match_efficiency_pct':'Match Eff %',
        'avg_cap_eff':'Avg Cap Eff %','total_parts':'Parts','production_hrs':'Prod Hrs',
        'tools':'Tools','machines_used':'Machines',
    }) if not mer_df.empty else pd.DataFrame(), title="Match Efficiency Rate by Supplier")

    # ── Sheet 3: Supplier Scorecard ───────────────────────────────────────────
    ws_sc = wb.create_sheet("Supplier Scorecard")
    _write_sheet(ws_sc, scorecard if not scorecard.empty else pd.DataFrame(),
                 title="Supplier Scorecard")

    # ── Sheet 4: Optimal Pairings ─────────────────────────────────────────────
    ws_rec = wb.create_sheet("Optimal Pairings")
    _write_sheet(ws_rec, recs_df[[c for c in [
        'machine_id','best_tool','best_supplier','best_cap_eff','best_parts_per_hr',
        'worst_tool','worst_supplier','worst_cap_eff','tools_compared','cap_eff_spread',
        'cap_eff_gain'] if c in recs_df.columns]] if not recs_df.empty else pd.DataFrame(),
        title="Optimal Machine-Tool Pairings")

    # ── Sheet 5: Machine Rankings ─────────────────────────────────────────────
    ws_rk = wb.create_sheet("Machine Rankings")
    _write_sheet(ws_rk, rankings_df if not rankings_df.empty else pd.DataFrame(),
                 title="Machine Tool Rankings (All Sessions)")

    # ── Sheet 6: All Sessions ─────────────────────────────────────────────────
    ws_all = wb.create_sheet("All Sessions")
    export_cols = [c for c in [
        'tool_id','supplier_id','machine_id','runs','production_hrs','total_parts',
        'cap_efficiency_pct','stability_pct','efficiency_pct','fit_score',
        'avg_ct_sec','mtbf_min','mttr_min','stop_count','downtime_hrs',
        'slow_loss_parts','fast_gain_parts'
    ] if c in fit_df.columns]
    _write_sheet(ws_all, fit_df[export_cols].round(2) if not fit_df.empty else pd.DataFrame(),
                 title="All Machine-Tool Sessions")

    wb.save(output)
    return output.getvalue()


# ==============================================================================
# --- PART ANALYSIS ENGINE ---
# ==============================================================================

def run_part_analysis(
    fit_df: pd.DataFrame,
    df_processed: pd.DataFrame,
    tool_ids: list,
    part_name: str,
    copy_map: dict = None,
) -> list:
    """
    Rules-based analysis engine for a part (one or more copy tools).
    Returns a list of insight dicts:
        {
            'rule':     str,   # rule name
            'severity': str,   # 'high' | 'medium' | 'info'
            'title':    str,   # one-line summary
            'detail':   str,   # explanation with numbers
        }
    """
    insights = []
    if fit_df.empty or not tool_ids:
        return insights

    part_fit = fit_df[fit_df['tool_id'].isin(tool_ids)].copy()
    if part_fit.empty:
        return insights

    # ── 1. Consistency flag ───────────────────────────────────────────────────
    for tid, tgrp in part_fit.groupby('tool_id'):
        if len(tgrp) < 2:
            continue
        cap_range = tgrp['cap_efficiency_pct'].max() - tgrp['cap_efficiency_pct'].min()
        best_m  = tgrp.loc[tgrp['cap_efficiency_pct'].idxmax(), 'machine_id']
        worst_m = tgrp.loc[tgrp['cap_efficiency_pct'].idxmin(), 'machine_id']
        best_v  = tgrp['cap_efficiency_pct'].max()
        worst_v = tgrp['cap_efficiency_pct'].min()
        if cap_range > 15:
            insights.append({
                'rule': 'consistency',
                'severity': 'high',
                'tool_id': tid, 'machine_id': best_m,
                'title': f"{tid} — High machine-dependency ({cap_range:.0f} pp spread)",
                'detail': (
                    f"{tid} performs very differently depending on which machine it runs on. "
                    f"Best: {best_m} at {best_v:.0f}%, worst: {worst_m} at {worst_v:.0f}%. "
                    f"A {cap_range:.0f} pp spread suggests the tool is sensitive to press conditions. "
                    f"Prioritise running on {best_m}."
                ),
            })
        elif cap_range > 7:
            insights.append({
                'rule': 'consistency',
                'severity': 'medium',
                'tool_id': tid, 'machine_id': best_m,
                'title': f"{tid} — Moderate machine sensitivity ({cap_range:.0f} pp spread)",
                'detail': (
                    f"{tid} shows a {cap_range:.0f} pp range across machines "
                    f"({best_m}: {best_v:.0f}% → {worst_m}: {worst_v:.0f}%). "
                    f"Worth prioritising {best_m} for scheduling."
                ),
            })

    # ── 2. Shift degradation ──────────────────────────────────────────────────
    if (not df_processed.empty
            and 'session_period' in df_processed.columns
            and 'machine_id' in df_processed.columns):

        part_proc = df_processed[df_processed['tool_id'].isin(tool_ids)].copy()

        def _cap_eff_for_group(grp):
            prod = grp[grp['stop_flag'] == 0]
            dur  = (grp['shot_time'].max() - grp['shot_time'].min()).total_seconds()
            if dur <= 0: return None
            rct  = float(grp['approved_ct_for_run'].iloc[0]) if 'approved_ct_for_run' in grp.columns else float(grp['approved_ct'].iloc[0])
            rcav = float(grp['working_cavities'].max()) if 'working_cavities' in grp.columns else 1.0
            opt  = (dur / rct) * rcav if rct > 0 else 0
            act  = float(prod['working_cavities'].sum()) if 'working_cavities' in prod.columns else float(len(prod))
            return round(act / opt * 100, 1) if opt > 0 else None

        if not part_proc.empty:
            shift_effs = {}
            for (tid, mid, period), grp in part_proc.groupby(['tool_id','machine_id','session_period']):
                v = _cap_eff_for_group(grp)
                if v is not None:
                    shift_effs.setdefault((tid, mid), {})[period] = v

            for (tid, mid), periods in shift_effs.items():
                if len(periods) < 2:
                    continue
                vals = list(periods.values())
                max_v = max(vals); min_v = min(vals)
                max_s = max(periods, key=periods.get)
                min_s = min(periods, key=periods.get)
                drop = max_v - min_v
                if drop > 10:
                    insights.append({
                        'rule': 'shift_degradation',
                        'severity': 'high',
                        'tool_id': tid, 'machine_id': mid,
                        'shift_filter': min_s,
                        'title': f"{tid} / {mid} — Shift degradation detected ({drop:.0f} pp)",
                        'detail': (
                            f"{tid} on {mid} drops {drop:.0f} pp between best and worst shift. "
                            f"Peak: {max_s} at {max_v:.0f}%, lowest: {min_s} at {min_v:.0f}%. "
                            f"Possible causes: thermal drift, operator differences, or end-of-shift fatigue. "
                            f"Investigate conditions during {min_s}."
                        ),
                    })
                elif drop > 5:
                    insights.append({
                        'rule': 'shift_degradation',
                        'severity': 'medium',
                        'tool_id': tid, 'machine_id': mid,
                        'shift_filter': min_s,
                        'title': f"{tid} / {mid} — Shift variation ({drop:.0f} pp)",
                        'detail': (
                            f"{tid} on {mid} shows {drop:.0f} pp variation across shifts "
                            f"({max_s}: {max_v:.0f}% → {min_s}: {min_v:.0f}%). "
                            f"Worthwhile investigating {min_s} conditions."
                        ),
                    })

    # ── 3. Underutilised pairing ──────────────────────────────────────────────
    for tid, tgrp in part_fit.groupby('tool_id'):
        if len(tgrp) > 1:
            continue  # already tested on multiple machines
        tested_m = tgrp.iloc[0]['machine_id']
        tested_v = tgrp.iloc[0]['cap_efficiency_pct']
        sup      = tgrp.iloc[0]['supplier_id'] if 'supplier_id' in tgrp.columns else None

        # Find other machines where similar tools (same supplier) perform better
        if sup:
            similar = fit_df[
                (fit_df['supplier_id'] == sup) &
                (~fit_df['tool_id'].isin(tool_ids))
            ]
            better = similar[similar['cap_efficiency_pct'] > tested_v + 5]
            if not better.empty:
                best_alt = better.loc[better['cap_efficiency_pct'].idxmax()]
                insights.append({
                    'rule': 'underutilised',
                    'severity': 'medium',
                    'tool_id': tid, 'machine_id': tested_m,
                    'title': f"{tid} — Only tested on one machine",
                    'detail': (
                        f"{tid} has only ever run on {tested_m} ({tested_v:.0f}% cap eff). "
                        f"Other {sup} tools perform better on {best_alt['machine_id']} "
                        f"({best_alt['cap_efficiency_pct']:.0f}% cap eff). "
                        f"Recommend trialling {tid} on {best_alt['machine_id']}."
                    ),
                })

    # ── 4. Best window ────────────────────────────────────────────────────────
    if not df_processed.empty and 'run_id' in df_processed.columns and 'machine_id' in df_processed.columns:
        part_proc = df_processed[df_processed['tool_id'].isin(tool_ids)].copy()
        if not part_proc.empty:
            best_run_eff = -1; best_run_info = None
            for (tid, mid, run_id), rg in part_proc.groupby(['tool_id','machine_id','run_id']):
                prod = rg[rg['stop_flag'] == 0]
                dur  = (rg['shot_time'].max() - rg['shot_time'].min()).total_seconds()
                if dur < 600: continue  # skip runs < 10 min
                rct  = float(rg['approved_ct_for_run'].iloc[0]) if 'approved_ct_for_run' in rg.columns else float(rg['approved_ct'].iloc[0])
                rcav = float(rg['working_cavities'].max()) if 'working_cavities' in rg.columns else 1.0
                opt  = (dur / rct) * rcav if rct > 0 else 0
                act  = float(prod['working_cavities'].sum()) if 'working_cavities' in prod.columns else float(len(prod))
                eff  = act / opt * 100 if opt > 0 else 0
                if eff > best_run_eff:
                    best_run_eff  = eff
                    best_run_info = {
                        'tool_id':      tid,
                        'machine_id':   mid,
                        'start':        rg['shot_time'].min().strftime('%d %b %Y %H:%M'),
                        'end':          rg['shot_time'].max().strftime('%d %b %Y %H:%M'),
                        'ts_start':     rg['shot_time'].min(),
                        'ts_end':       rg['shot_time'].max(),
                        'eff':          round(eff, 0),
                        'hrs':          round(dur / 3600, 1),
                        'period':       rg['session_period'].iloc[0] if 'session_period' in rg.columns else '—',
                    }
            if best_run_info:
                r = best_run_info
                insights.append({
                    'rule': 'best_window',
                    'severity': 'info',
                    'tool_id': r['tool_id'], 'machine_id': r['machine_id'],
                    'date_from': r['ts_start'], 'date_to': r['ts_end'],
                    'title': f"Peak performance: {r['tool_id']} on {r['machine_id']} — {r['eff']:.0f}% cap eff",
                    'detail': (
                        f"Best recorded production window: {r['start']} → {r['end']} "
                        f"({r['hrs']:.1f} hrs, {r['period']}). "
                        f"Cap efficiency: {r['eff']:.0f}%. "
                        f"Use this as the benchmark for what this part is capable of."
                    ),
                })

    # ── 5. Sibling divergence ─────────────────────────────────────────────────
    if len(tool_ids) > 1:
        tool_avgs = part_fit.groupby('tool_id')['cap_efficiency_pct'].mean()
        if len(tool_avgs) > 1:
            best_tool  = tool_avgs.idxmax()
            worst_tool = tool_avgs.idxmin()
            gap = tool_avgs.max() - tool_avgs.min()
            if gap > 10:
                insights.append({
                    'rule': 'sibling_divergence',
                    'severity': 'high',
                    'tool_id': worst_tool, 'machine_id': None,
                    'title': f"Copy tool divergence — {gap:.0f} pp gap between {best_tool} and {worst_tool}",
                    'detail': (
                        f"{best_tool} averages {tool_avgs[best_tool]:.0f}% cap efficiency across all machines, "
                        f"while {worst_tool} averages only {tool_avgs[worst_tool]:.0f}%. "
                        f"These are copy tools of the same part — a {gap:.0f} pp gap suggests "
                        f"tooling wear, damage, or maintenance differences. "
                        f"Recommend physical inspection of {worst_tool}."
                    ),
                })
            elif gap > 5:
                insights.append({
                    'rule': 'sibling_divergence',
                    'severity': 'medium',
                    'tool_id': worst_tool, 'machine_id': None,
                    'title': f"Copy tool gap — {gap:.0f} pp between {best_tool} and {worst_tool}",
                    'detail': (
                        f"{best_tool} ({tool_avgs[best_tool]:.0f}%) is outperforming "
                        f"{worst_tool} ({tool_avgs[worst_tool]:.0f}%) by {gap:.0f} pp on average. "
                        f"Monitor {worst_tool} for wear indicators."
                    ),
                })

    # ── 6. Stagnation / decline ───────────────────────────────────────────────
    if not df_processed.empty and 'run_id' in df_processed.columns and 'machine_id' in df_processed.columns:
        part_proc = df_processed[df_processed['tool_id'].isin(tool_ids)].copy()
        if not part_proc.empty:
            for (tid, mid), grp in part_proc.groupby(['tool_id','machine_id']):
                run_effs = []
                for run_id, rg in grp.groupby('run_id'):
                    prod = rg[rg['stop_flag'] == 0]
                    dur  = (rg['shot_time'].max() - rg['shot_time'].min()).total_seconds()
                    if dur < 600: continue
                    rct  = float(rg['approved_ct_for_run'].iloc[0]) if 'approved_ct_for_run' in rg.columns else float(rg['approved_ct'].iloc[0])
                    rcav = float(rg['working_cavities'].max()) if 'working_cavities' in rg.columns else 1.0
                    opt  = (dur / rct) * rcav if rct > 0 else 0
                    act  = float(prod['working_cavities'].sum()) if 'working_cavities' in prod.columns else float(len(prod))
                    eff  = act / opt * 100 if opt > 0 else 0
                    run_effs.append((rg['shot_time'].min(), eff))

                if len(run_effs) < 4:
                    continue

                run_effs.sort(key=lambda x: x[0])
                effs = [e for _, e in run_effs]
                first_half  = np.mean(effs[:len(effs)//2])
                second_half = np.mean(effs[len(effs)//2:])
                trend = second_half - first_half

                if trend < -8:
                    insights.append({
                        'rule': 'stagnation',
                        'severity': 'high',
                        'tool_id': tid, 'machine_id': mid,
                        'title': f"{tid} / {mid} — Declining performance ({trend:.0f} pp trend)",
                        'detail': (
                            f"{tid} on {mid} is declining over time. "
                            f"First half of runs averaged {first_half:.0f}% cap eff, "
                            f"second half averaged {second_half:.0f}% — a {abs(trend):.0f} pp drop. "
                            f"Across {len(effs)} runs. "
                            f"Likely tooling wear or machine degradation. Maintenance recommended."
                        ),
                    })
                elif trend < -4:
                    insights.append({
                        'rule': 'stagnation',
                        'severity': 'medium',
                        'tool_id': tid, 'machine_id': mid,
                        'title': f"{tid} / {mid} — Gradual decline ({trend:.0f} pp trend)",
                        'detail': (
                            f"{tid} on {mid} shows a gradual downward trend across {len(effs)} runs "
                            f"({first_half:.0f}% → {second_half:.0f}%). Monitor closely."
                        ),
                    })
                elif -2 <= trend <= 2 and len(effs) >= 6:
                    insights.append({
                        'rule': 'stagnation',
                        'severity': 'info',
                        'tool_id': tid, 'machine_id': mid,
                        'title': f"{tid} / {mid} — No improvement across {len(effs)} runs",
                        'detail': (
                            f"{tid} on {mid} has shown no meaningful change across {len(effs)} runs "
                            f"(avg: {np.mean(effs):.0f}%). "
                            f"If below target, check process parameters — the pairing appears stable but sub-optimal."
                        ),
                    })

    # Sort: high → medium → info
    order = {'high': 0, 'medium': 1, 'info': 2}
    insights.sort(key=lambda x: order.get(x['severity'], 3))
    return insights
